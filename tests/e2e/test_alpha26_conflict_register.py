"""Alpha-26: conflict register peer surface — JSON + xlsx endpoints."""

from __future__ import annotations

import io
from pathlib import Path

import pytest


def test_conflict_register_returns_all_statuses_by_default(
    project_with_conflicts_mixed_status, api_client,
):
    slug = project_with_conflicts_mixed_status
    res = api_client.get(f"/api/projects/{slug}/conflict-register")
    assert res.status_code == 200
    body = res.json()
    assert "items" in body
    assert "counts" in body
    assert body["counts"]["all"] == len(body["items"])
    assert body["counts"]["pending"] >= 1
    assert body["counts"]["resolved"] >= 1


def test_conflict_register_filters_by_status(
    project_with_conflicts_mixed_status, api_client,
):
    slug = project_with_conflicts_mixed_status
    res_pending = api_client.get(
        f"/api/projects/{slug}/conflict-register?status=pending"
    )
    body_pending = res_pending.json()
    for item in body_pending["items"]:
        assert item["status"] == "pending"

    res_resolved = api_client.get(
        f"/api/projects/{slug}/conflict-register?status=resolved"
    )
    body_resolved = res_resolved.json()
    for item in body_resolved["items"]:
        assert item["status"].startswith("resolved_")


def test_conflict_register_includes_resolved_at_for_resolved_rows(
    project_with_conflicts_mixed_status, api_client,
):
    slug = project_with_conflicts_mixed_status
    res = api_client.get(f"/api/projects/{slug}/conflict-register?status=resolved")
    body = res.json()
    for item in body["items"]:
        assert item.get("resolved_at"), "resolved row missing resolved_at"


def test_conflict_register_xlsx_round_trip(
    project_with_conflicts_mixed_status, api_client, tmp_path,
):
    from openpyxl import load_workbook

    slug = project_with_conflicts_mixed_status
    res = api_client.get(f"/api/projects/{slug}/conflict-register.xlsx")
    assert res.status_code == 200
    out = tmp_path / "conflicts.xlsx"
    out.write_bytes(res.content)
    wb = load_workbook(out)
    assert "Conflicts" in wb.sheetnames
    ws = wb["Conflicts"]
    headers = [c.value for c in ws[1]]
    for required in [
        "source A", "value A", "source B", "value B",
        "kind", "most-onerous reasoning", "status", "resolution", "resolved at",
    ]:
        assert required in headers, f"missing column: {required}; got {headers}"


def test_conflict_register_xlsx_resolved_row_has_acceptA_resolution(
    project_with_conflicts_mixed_status, api_client, tmp_path,
):
    from openpyxl import load_workbook
    slug = project_with_conflicts_mixed_status
    res = api_client.get(f"/api/projects/{slug}/conflict-register.xlsx")
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb["Conflicts"]
    headers = [c.value for c in ws[1]]
    res_col = headers.index("resolution") + 1
    status_col = headers.index("status") + 1
    found_acceptA = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=status_col).value == "resolved_accept_a":
            assert ws.cell(row=row, column=res_col).value == "Accept A"
            found_acceptA = True
            break
    assert found_acceptA, "fixture should have at least one resolved_accept_a row"
