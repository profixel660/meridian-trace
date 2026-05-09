"""End-to-end tests for the extraction orchestrator using a stubbed LLM.

Covers:
* chunk-level resume gating (v5)
* atomic persist + EJS finalisation
* the three-outcome (inside / borderline / outside) split
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest

from meridian.extract.orchestrator import (
    create_job,
    finish_job,
    run_extraction_for_source,
    run_job_over_sources,
)


def test_chunk_resume_skips_extracted_chunks(
    fresh_project_with_sample_doc: tuple[str, Path, sqlite3.Connection, str],
    mock_llm_client,
) -> None:
    """Pre-mark every chunk for the source as ``extracted``; triage must
    see them as already-terminal and emit zero new triage decisions."""
    _name, _db, conn, source_id = fresh_project_with_sample_doc

    chunk_ids = [
        r["id"]
        for r in conn.execute(
            "SELECT id FROM source_document_chunk WHERE source_id = ?",
            (source_id,),
        ).fetchall()
    ]
    assert chunk_ids, "fixture must produce at least one chunk"

    # Pre-mark every chunk as 'extracted' with keep=true so the resume gate
    # in run_triage_for_source treats them as terminal.
    with conn:
        for cid in chunk_ids:
            conn.execute(
                "UPDATE source_document_chunk "
                "SET extraction_status = 'extracted', "
                "    triage_marked_for_extraction = 1, "
                "    extraction_finished_at = '2026-01-01T00:00:00Z' "
                "WHERE id = ?",
                (cid,),
            )

    from meridian.extract.triage import run_triage_for_source

    job_id = create_job(conn)
    try:
        result = run_triage_for_source(conn, source_id=source_id, job_id=job_id)
        # Every chunk was already terminal → no LLM triage call should fire.
        assert result.triaged == 0
        assert result.kept == len(chunk_ids)
        triage_calls = [r for r in mock_llm_client.recorded if r[0] == "triage"]
        assert triage_calls == [], f"unexpected triage call(s): {triage_calls}"
    finally:
        finish_job(conn, job_id=job_id)


def test_ejs_finalisation_is_transactional(
    fresh_project_with_sample_doc: tuple[str, Path, sqlite3.Connection, str],
    mock_llm_client,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Force ``persist_parsed_output`` to raise mid-flight. Afterwards the
    EJS row must be ``failed`` and NO deliverable rows must exist for that
    source — the orchestrator wraps persist + state-flip in one transaction
    so a partial commit is impossible."""
    _name, _db, conn, source_id = fresh_project_with_sample_doc

    job_id = create_job(conn)
    try:
        # Sabotage the persist call AFTER the LLM returns successfully so we
        # exercise the atomic-boundary code path specifically (round-9 §6 B).
        def _boom(*args, **kwargs):  # noqa: ANN001, ANN002
            raise RuntimeError("synthetic persist failure")

        monkeypatch.setattr(
            "meridian.extract.orchestrator.persist_parsed_output", _boom
        )

        out = run_extraction_for_source(conn, job_id=job_id, source_id=source_id)
        assert out.error is not None
        assert "synthetic persist failure" in out.error

        # No deliverable rows — the persist transaction rolled back.
        n_deliv = conn.execute(
            "SELECT COUNT(*) FROM deliverable WHERE source_id = ?", (source_id,)
        ).fetchone()[0]
        assert n_deliv == 0

        # EJS recorded as 'failed' (NOT left in 'extracting' which would be
        # the torn-write the test guards against).
        ejs_status = conn.execute(
            "SELECT status FROM extraction_job_source "
            "WHERE job_id = ? AND source_id = ?",
            (job_id, source_id),
        ).fetchone()["status"]
        assert ejs_status == "failed", f"expected 'failed', got {ejs_status!r}"
    finally:
        finish_job(conn, job_id=job_id)


def test_three_outcome_classification(
    fresh_project_with_sample_doc: tuple[str, Path, sqlite3.Connection, str],
    mock_llm_client,
) -> None:
    """Feed the LLM stub a payload with one INSIDE, one BORDERLINE, one OUTSIDE
    candidate; assert each lands in the right table with the right gate_outcome.

    Borderline is signalled via the ``definition_borderline`` flag on a
    deliverable row (per ``persist._decide_auto_route``); OUTSIDE candidates
    arrive in the ``audit`` array.
    """
    _name, _db, conn, _src_id = fresh_project_with_sample_doc

    mock_llm_client.responses["extract_text_spec"] = {
        "deliverables": [
            {
                "trade": "Mechanical",
                "service": "HVAC",
                "category": "design",
                "deliverables_summary": "INSIDE: clear deliverable.",
                "confidence": "high",
                "flags": [],
                "source_ref": "p.1",
                "applicable_standards": [],
            },
            {
                "trade": "Mechanical",
                "service": "HVAC",
                "category": "design",
                "deliverables_summary": "BORDERLINE: ambiguous wording.",
                "confidence": "medium",
                "flags": ["definition_borderline"],
                "source_ref": "p.2",
                "applicable_standards": [],
            },
        ],
        "audit": [
            {
                "candidate_text": "OUTSIDE: marketing copy, not a deliverable.",
                "rejection_reason": "non-deliverable language",
                "source_ref": "p.3",
            }
        ],
        "questions": [],
    }

    result = run_job_over_sources(conn, source_ids=[_src_id])
    assert result.sources, "orchestrator returned no source results"
    src_result = result.sources[0]
    assert src_result.error is None, src_result.error

    counts = src_result.counts
    assert counts["inside"] == 1
    assert counts["borderline"] == 1
    assert counts["outside"] == 1

    # Verify the rows landed in the right tables with the right gate.
    deliv_rows = conn.execute(
        "SELECT gate_outcome FROM deliverable WHERE source_id = ?", (_src_id,)
    ).fetchall()
    gates = sorted(r["gate_outcome"] for r in deliv_rows)
    assert gates == ["borderline", "inside"]

    audit_rows = conn.execute(
        "SELECT candidate_text FROM audit_record WHERE source_id = ?", (_src_id,)
    ).fetchall()
    assert len(audit_rows) == 1
    assert "OUTSIDE" in audit_rows[0]["candidate_text"]


def test_extraction_records_zero_real_llm_calls(
    fresh_project_with_sample_doc: tuple[str, Path, sqlite3.Connection, str],
    mock_llm_client,
) -> None:
    """Sanity check: a full extraction run must not have hit the network.

    Every ``llm_call`` row should have provider=='stub' (set by the mock).
    A real call would have ``provider='anthropic'`` (the default route) and
    would have crashed earlier on the missing API key — this guards against
    a future regression where the monkeypatch silently fails to apply.
    """
    _name, _db, conn, src_id = fresh_project_with_sample_doc

    run_job_over_sources(conn, source_ids=[src_id])

    providers = {
        r["provider"]
        for r in conn.execute("SELECT DISTINCT provider FROM llm_call").fetchall()
    }
    assert providers == {"stub"}, f"unexpected providers in llm_call: {providers}"


def test_alpha16_full_loop_extraction_to_excel_export(
    fresh_project_with_sample_doc: tuple[str, Path, sqlite3.Connection, str],
    mock_llm_client,
    tmp_path: Path,
) -> None:
    """Alpha-16 core-validation gate.

    The full deliverables-extraction loop the user is trying to validate:
    real document already imported (via fixture) -> orchestrator runs
    over it with mock LLM producing 2 INSIDE deliverables -> deliverable
    rows land in the project DB -> export_to_xlsx produces a real .xlsx
    file with both rows on the Master sheet.

    Until alpha-16, no test exercised export_to_xlsx end-to-end. The
    extraction-side tests in this file cover the pipeline up to
    persistence; this one closes the gap to the actual product output
    (the per-trade Excel register the user receives)."""
    from openpyxl import load_workbook

    from meridian.export.excel import export_to_xlsx

    _name, _db, conn, src_id = fresh_project_with_sample_doc

    # Mock LLM produces two INSIDE deliverables on different trades so
    # the pivot sheets exercise the group-by paths too.
    mock_llm_client.responses["extract_text_spec"] = {
        "deliverables": [
            {
                "trade": "Mechanical",
                "service": "HVAC",
                "category": "design",
                "deliverables_summary": "Provide ductwork design drawings to AS 4254.",
                "confidence": "high",
                "flags": [],
                "source_ref": "p.1",
                "applicable_standards": ["AS 4254"],
            },
            {
                "trade": "Electrical",
                "service": "Power",
                "category": "supply",
                "deliverables_summary": "Supply and install MSB to AS/NZS 3000.",
                "confidence": "high",
                "flags": [],
                "source_ref": "p.2",
                "applicable_standards": ["AS/NZS 3000"],
            },
        ],
        "audit": [],
        "questions": [],
    }

    result = run_job_over_sources(conn, source_ids=[src_id])
    assert result.sources and result.sources[0].error is None
    assert result.sources[0].counts["inside"] == 2

    # Two deliverables landed; the orchestrator's auto-route may
    # quarantine a subset (e.g. for new-taxonomy review). Promote any
    # quarantined rows to 'user_accepted' so both reach the master
    # register — that mirrors what the operator would do clicking
    # "Accept" in the Quarantine queue. This test exercises the
    # extract -> quarantine-decision -> master -> export path the
    # alpha-16 strip preserves.
    deliv_rows = conn.execute(
        "SELECT id, status FROM deliverable WHERE source_id = ?",
        (src_id,),
    ).fetchall()
    assert len(deliv_rows) == 2, f"expected 2 deliverables persisted, got {len(deliv_rows)}"
    with conn:
        for row in deliv_rows:
            if row["status"] == "quarantined":
                conn.execute(
                    "UPDATE deliverable SET status = 'user_accepted' WHERE id = ?",
                    (row["id"],),
                )
    master_count = conn.execute(
        "SELECT COUNT(*) AS n FROM v_master_register"
    ).fetchone()["n"]
    assert master_count == 2, f"expected 2 deliverables on the master register, got {master_count}"

    # Now exercise the actual product output: the Excel file.
    out_path = tmp_path / "alpha16-export.xlsx"
    export_to_xlsx(conn, output_path=out_path)
    assert out_path.exists(), f"export did not produce a file at {out_path}"
    assert out_path.stat().st_size > 0, "exported xlsx is empty"

    wb = load_workbook(filename=str(out_path), read_only=True, data_only=True)
    try:
        # Master sheet should be the active one and contain both rows.
        assert "Master" in wb.sheetnames, wb.sheetnames
        master = wb["Master"]
        # First row is the header. Following rows are deliverables.
        rows = list(master.iter_rows(values_only=True))
        assert len(rows) >= 3, f"expected header + at least 2 data rows, got {len(rows)}"
        body = [r for r in rows[1:] if any(cell is not None for cell in r)]
        # The deliverables_summary column should contain our two strings.
        flat = " ".join(
            str(cell) for row in body for cell in row if cell is not None
        )
        assert "ductwork" in flat.lower(), f"missing Mechanical deliverable text: {flat[:500]}"
        assert "msb" in flat.lower(), f"missing Electrical deliverable text: {flat[:500]}"

        # Pivot sheets must exist and not be empty.
        for pivot in ("By Trade", "By Service", "By Category"):
            assert pivot in wb.sheetnames, f"missing pivot sheet {pivot!r}; got {wb.sheetnames}"
            pivot_rows = list(wb[pivot].iter_rows(values_only=True))
            assert len(pivot_rows) >= 2, f"pivot {pivot!r} has no data rows"
    finally:
        wb.close()


def test_run_job_over_sources_invokes_on_source_complete(
    project_with_two_sources, mock_llm_client, monkeypatch,
):
    """The on_source_complete callback fires once per source (in order)."""
    conn, source_ids, source_filenames = project_with_two_sources
    seen: list[tuple[str, str]] = []

    def cb(source_id: str, filename: str) -> None:
        seen.append((source_id, filename))

    run_job_over_sources(
        conn, source_ids=source_ids, on_source_complete=cb,
    )
    assert len(seen) == len(source_ids)
    assert [sid for sid, _ in seen] == source_ids
    # Filenames are passed verbatim — no normalisation
    assert all(fn for _, fn in seen)
