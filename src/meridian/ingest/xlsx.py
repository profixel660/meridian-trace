"""XLSX ingestion via openpyxl.

For BOD-style sheets the chunk granularity is one chunk per non-header row,
carrying the row's cell values keyed by column letter. The sheet's header row
(row 1) is captured in the chunk metadata so downstream extractors can use
named-column access.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from meridian.ingest._types import _Chunk, _Extracted


def _row_to_text(headers: list[str | None], values: list[Any]) -> str:
    """Render a row as 'Header: value' lines for inclusion in the LLM prompt."""
    parts: list[str] = []
    for header, value in zip(headers, values, strict=False):
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        label = header if header else "(unlabelled)"
        parts.append(f"{label}: {value}")
    return "\n".join(parts)


def extract(path: Path) -> _Extracted:
    """Extract per-sheet/per-row chunks suitable for the BOD prompt."""
    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    chunks: list[_Chunk] = []
    sheet_summaries: list[str] = []
    metadata: dict[str, Any] = {"sheets": []}

    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        first = next(rows_iter, None)
        headers: list[str | None] = (
            [str(c).strip() if c is not None else None for c in first] if first else []
        )

        sheet_chunks = 0
        sheet_lines: list[str] = []
        sheet_lines.append(f"### Sheet: {sheet.title}")
        if headers:
            sheet_lines.append("Headers: " + " | ".join(h or "(blank)" for h in headers))

        for row_index, row in enumerate(rows_iter, start=2):
            values = list(row)
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            row_text = _row_to_text(headers, values)
            chunks.append(
                _Chunk(
                    chunk_kind="xlsx_row",
                    locator={
                        "sheet": sheet.title,
                        "row": row_index,
                        "headers": headers,
                    },
                    text=row_text,
                )
            )
            sheet_lines.append(f"-- Row {row_index} --\n{row_text}")
            sheet_chunks += 1

        metadata["sheets"].append(
            {
                "name": sheet.title,
                "headers": headers,
                "row_count": sheet_chunks,
            }
        )
        sheet_summaries.append("\n".join(sheet_lines))

    workbook.close()

    full_text = "\n\n".join(sheet_summaries)
    return _Extracted(
        extraction_method="xlsx_parsed",
        text=full_text,
        chunks=chunks,
        metadata=metadata,
    )


__all__ = ["extract"]
