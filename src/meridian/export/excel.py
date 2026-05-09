"""Excel export — Master + By-Trade / By-Service / By-Category pivots.

Per CONTEXT.md §15: v1 is export-only. Excel regenerates from SQLite on demand.
The structured `source_ref` is rendered to its `rendered` field; raw JSON is not exposed.
"""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Master sheet column order — mirrors CONTEXT.md §4.
_MASTER_COLUMNS = [
    "id",
    "index",
    "source_document",
    "source_ref",
    "trade",
    "service",
    "category",
    "applicable_standards",
    "document_state",
    "document_class",
    "confidence",
    "flags",
    "conflict_summary",
    "deliverables_summary",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format_conflict_summary(
    conn: sqlite3.Connection, flag_context_raw: str | None,
) -> str:
    """Return newline-joined pending conflict summaries for a deliverable row.

    Reads ``flag_context`` JSON, extracts conflict_ids, queries the conflict
    table for rows that are still pending, and returns the LLM's
    ``most_onerous_reasoning`` verbatim with a ``[<kind>]`` prefix per row.
    Resolved conflicts (any resolved_* status) are intentionally excluded so
    the column only surfaces unresolved items requiring PM attention.
    """
    if not flag_context_raw:
        return ""
    try:
        ctx = json.loads(flag_context_raw)
    except (json.JSONDecodeError, TypeError):
        return ""
    if not isinstance(ctx, dict):
        return ""
    conflict_ids = [
        v.get("conflict_id")
        for v in ctx.values()
        if isinstance(v, dict) and v.get("conflict_id")
    ]
    if not conflict_ids:
        return ""
    placeholders = ",".join("?" * len(conflict_ids))
    rows = conn.execute(
        f"SELECT kind, most_onerous_reasoning FROM conflict "
        f"WHERE id IN ({placeholders}) AND status = 'pending'",
        conflict_ids,
    ).fetchall()
    return "\n".join(
        f"[{r['kind']}] {r['most_onerous_reasoning']}"
        for r in rows
        if r["most_onerous_reasoning"]
    )


def _row_payload(row: sqlite3.Row, conn: sqlite3.Connection) -> dict[str, str]:
    """Flatten a v_master_register row into the Excel column shape."""
    source_ref = row["source_ref"]
    if isinstance(source_ref, str):
        try:
            source_ref_obj = json.loads(source_ref)
        except json.JSONDecodeError:
            source_ref_obj = {"rendered": source_ref}
    else:
        source_ref_obj = source_ref or {}

    standards = row["applicable_standards"] or "[]"
    try:
        standards_list = json.loads(standards) if isinstance(standards, str) else standards
    except json.JSONDecodeError:
        standards_list = []

    flags = row["flags"] or "[]"
    try:
        flags_list = json.loads(flags) if isinstance(flags, str) else flags
    except json.JSONDecodeError:
        flags_list = []

    flag_context_raw = row["flag_context"] if "flag_context" in row.keys() else None
    return {
        "id": row["id"],
        "source_document": row["source_filename"],
        "source_ref": source_ref_obj.get("rendered") or "",
        "trade": row["trade_value"] or "",
        "service": row["service_value"] or "",
        "category": row["category_value"] or "",
        "applicable_standards": ", ".join(str(s) for s in standards_list),
        "document_state": row["document_state"] or "",
        "document_class": row["document_class"] or "",
        "confidence": row["confidence"] or "",
        "flags": ", ".join(str(f) for f in flags_list),
        "conflict_summary": _format_conflict_summary(conn, flag_context_raw),
        "deliverables_summary": row["deliverables_summary"] or "",
    }


def _select_master(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT
            d.id,
            d.source_ref,
            d.applicable_standards,
            d.confidence,
            d.flags,
            d.flag_context,
            d.deliverables_summary,
            d.created_at,
            sd.filename     AS source_filename,
            sd.document_class,
            sd.document_state,
            tt.value        AS trade_value,
            st.value        AS service_value,
            ct.value        AS category_value
        FROM v_master_register d
        LEFT JOIN source_document   sd ON sd.id = d.source_id
        LEFT JOIN trade_taxonomy    tt ON tt.id = d.trade_id
        LEFT JOIN service_taxonomy  st ON st.id = d.service_id
        LEFT JOIN category_taxonomy ct ON ct.id = d.category_id
        ORDER BY sd.filename, d.created_at, d.id
        """
    ).fetchall()


def _write_master_sheet(
    workbook: Workbook, rows: list[sqlite3.Row], conn: sqlite3.Connection,
) -> None:
    sheet = workbook.active
    sheet.title = "Master"

    sheet.append(_MASTER_COLUMNS)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    for index, row in enumerate(rows, start=1):
        payload = _row_payload(row, conn)
        sheet.append(
            [
                payload["id"],
                index,
                payload["source_document"],
                payload["source_ref"],
                payload["trade"],
                payload["service"],
                payload["category"],
                payload["applicable_standards"],
                payload["document_state"],
                payload["document_class"],
                payload["confidence"],
                payload["flags"],
                payload["conflict_summary"],
                payload["deliverables_summary"],
            ]
        )

    # Apply wrap_text on the conflict_summary column so multi-conflict rows
    # render correctly in Excel.
    conflict_col_letter = get_column_letter(_MASTER_COLUMNS.index("conflict_summary") + 1)
    for row_idx in range(2, sheet.max_row + 1):
        sheet[f"{conflict_col_letter}{row_idx}"].alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    # Column widths — heuristic, no hyperlink yet.
    widths = {
        "id": 38,
        "index": 6,
        "source_document": 38,
        "source_ref": 30,
        "trade": 20,
        "service": 22,
        "category": 14,
        "applicable_standards": 30,
        "document_state": 14,
        "document_class": 22,
        "confidence": 10,
        "flags": 26,
        "conflict_summary": 60,
        "deliverables_summary": 80,
    }
    for col_index, name in enumerate(_MASTER_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]

    sheet.freeze_panes = "B2"


def _write_pivot_sheet(
    workbook: Workbook,
    *,
    title: str,
    group_field: str,
    rows: list[sqlite3.Row],
    conn: sqlite3.Connection,
) -> None:
    sheet = workbook.create_sheet(title=title)
    grouped: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for row in rows:
        payload = _row_payload(row, conn)
        key = payload[group_field] or "(unset)"
        grouped[key].append(row)

    sheet.append([title.replace("By ", ""), "Count", "Summary preview"])
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    sheet.row_dimensions[1].height = 22

    for key in sorted(grouped):
        bucket = grouped[key]
        preview = "; ".join(_row_payload(r, conn)["deliverables_summary"] for r in bucket[:3])
        sheet.append([key, len(bucket), preview])

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 100
    sheet.freeze_panes = "A2"


def export_to_xlsx(conn: sqlite3.Connection, *, output_path: Path) -> Path:
    """Write Master + 3 pivot sheets to `output_path`. Returns the path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _select_master(conn)

    workbook = Workbook()
    _write_master_sheet(workbook, rows, conn)
    _write_pivot_sheet(workbook, title="By Trade", group_field="trade", rows=rows, conn=conn)
    _write_pivot_sheet(workbook, title="By Service", group_field="service", rows=rows, conn=conn)
    _write_pivot_sheet(workbook, title="By Category", group_field="category", rows=rows, conn=conn)

    workbook.save(str(output_path))
    return output_path


__all__ = ["export_to_xlsx"]
