"""Alpha-26 conflict register Excel emitter.

Single-sheet workbook. Columns mirror the page table verbatim; reasoning
column renders the conflict-pass LLM's most_onerous_reasoning unchanged
(per the project memory verbatim posture).
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_COLUMNS = [
    "source A", "value A", "source B", "value B",
    "kind", "most-onerous reasoning", "status", "resolution", "resolved at",
]

_RESOLUTION_LABEL = {
    "pending":              "—",
    "resolved_accept_a":    "Accept A",
    "resolved_accept_b":    "Accept B",
    "resolved_reject_both": "Reject both",
    "resolved_hybrid":      "Hybrid",
    "superseded":           "—",
}


def _select_register_rows(
    conn: sqlite3.Connection, *, status_filter: str = "all",
) -> list[dict[str, object]]:
    """Build the register rows. status_filter ∈ {all, pending, resolved, superseded}."""
    where_clauses = {
        "all": "1=1",
        "pending": "c.status = 'pending'",
        "resolved": "c.status LIKE 'resolved_%'",
        "superseded": "c.status = 'superseded'",
    }
    where = where_clauses.get(status_filter, "1=1")

    conflict_rows = conn.execute(f"""
        SELECT id, kind, status, most_onerous_reasoning, created_at, resolved_at
        FROM conflict c
        WHERE {where}
        ORDER BY (c.status = 'pending') DESC, c.created_at DESC
    """).fetchall()

    rows: list[dict[str, object]] = []
    for c in conflict_rows:
        parties = conn.execute("""
            SELECT party_kind, party_id, party_position
            FROM conflict_party WHERE conflict_id = ?
            ORDER BY rowid
        """, (c["id"],)).fetchall()

        # Look up source filenames for each party
        sources = []
        for p in parties:
            if p["party_kind"] == "deliverable":
                src = conn.execute("""
                    SELECT sd.filename
                    FROM deliverable d
                    LEFT JOIN source_document sd ON sd.id = d.source_id
                    WHERE d.id = ?
                """, (p["party_id"],)).fetchone()
                sources.append(src["filename"] if src else "(deliverable)")
            elif p["party_kind"] == "audit":
                src = conn.execute("""
                    SELECT sd.filename
                    FROM audit_record a
                    LEFT JOIN source_document sd ON sd.id = a.source_id
                    WHERE a.id = ?
                """, (p["party_id"],)).fetchone()
                sources.append(src["filename"] if src else "(audit)")
            else:
                sources.append("?")

        # Pad to 2 parties for the table
        parties_list = list(parties)
        while len(parties_list) < 2:
            parties_list.append({"party_position": ""})
            sources.append("")

        rows.append({
            "source A": sources[0],
            "value A": parties_list[0]["party_position"] or "",
            "source B": sources[1],
            "value B": parties_list[1]["party_position"] or "",
            "kind": c["kind"],
            "most-onerous reasoning": c["most_onerous_reasoning"] or "",
            "status": c["status"],
            "resolution": _RESOLUTION_LABEL.get(c["status"], "—"),
            "resolved at": c["resolved_at"] or c["created_at"],
        })
    return rows


def export_conflict_register_xlsx(
    conn: sqlite3.Connection, *, output_path: Path,
) -> Path:
    """Write the project's conflict register to a single-sheet workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _select_register_rows(conn, status_filter="all")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conflicts"

    sheet.append(_COLUMNS)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    for row in rows:
        sheet.append([row[col] for col in _COLUMNS])

    # Wrap text on long-form columns
    wrap_cols = ["value A", "value B", "most-onerous reasoning"]
    for wc in wrap_cols:
        col_letter = get_column_letter(_COLUMNS.index(wc) + 1)
        for r in range(2, sheet.max_row + 1):
            sheet[f"{col_letter}{r}"].alignment = Alignment(
                wrap_text=True, vertical="top",
            )

    widths = {
        "source A": 30, "value A": 30, "source B": 30, "value B": 30,
        "kind": 18, "most-onerous reasoning": 60, "status": 18,
        "resolution": 14, "resolved at": 22,
    }
    for col_index, name in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]

    sheet.freeze_panes = "A2"
    workbook.save(str(output_path))
    return output_path


__all__ = ["export_conflict_register_xlsx"]
