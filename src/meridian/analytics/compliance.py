"""Compliance Traceability — pivot deliverables x applicable_standards.

Pure read-only over the per-project SQLite. Surfaces:
- which standards apply to which deliverables (matrix view)
- deliverables with no standards tied (potential compliance gap)
- over-cited / under-cited standards
- one-off standards (cited exactly once — worth a sanity check)

Strict citation rule (CONTEXT.md §4): we are NOT inferring standards from
document foreword / general references. The data is what was extracted; this
analytics module only counts and pivots — never invents.
"""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# Excel render style — kept consistent with src/meridian/export/excel.py.
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ONE_OFF_FILL = PatternFill("solid", fgColor="FFF3B0")  # yellow highlight

_SAMPLE_LIMIT = 50


# ─── Pydantic models ───────────────────────────────────────────────────────


class ComplianceMatrixCell(BaseModel):
    standard: str
    deliverable_count: int
    trades: list[str]
    services: list[str]


class ComplianceTraceabilityResult(BaseModel):
    project_name: str
    generated_at: str
    total_deliverables: int
    deliverables_with_standards: int
    deliverables_without_standards: int
    distinct_standards: int
    standards: list[ComplianceMatrixCell]
    deliverables_without_standards_sample: list[dict]
    one_off_standards: list[str]


# ─── Internal helpers ──────────────────────────────────────────────────────


def _parse_standards(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []
    if not isinstance(parsed, list):
        return []
    out: list[str] = []
    for item in parsed:
        if item is None:
            continue
        s = str(item).strip()
        if s:
            out.append(s)
    return out


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


# ─── Public API ───────────────────────────────────────────────────────────


def compute(conn: sqlite3.Connection) -> ComplianceTraceabilityResult:
    """Build the compliance pivot from the deliverable + taxonomy tables."""
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    project_row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    project_name = project_row["name"] if project_row else "(unknown)"

    rows = conn.execute(
        """
        SELECT
            d.id,
            d.applicable_standards,
            d.deliverables_summary,
            tt.value AS trade_value,
            st.value AS service_value,
            sd.filename AS source_filename
        FROM deliverable d
        LEFT JOIN trade_taxonomy   tt ON tt.id = d.trade_id
        LEFT JOIN service_taxonomy st ON st.id = d.service_id
        LEFT JOIN source_document  sd ON sd.id = d.source_id
        """
    ).fetchall()

    total_deliverables = len(rows)
    with_standards = 0
    without_standards_sample: list[dict] = []

    # standard -> set(deliverable_id), set(trade), set(service)
    by_standard_deliv: dict[str, set[str]] = defaultdict(set)
    by_standard_trades: dict[str, set[str]] = defaultdict(set)
    by_standard_services: dict[str, set[str]] = defaultdict(set)

    for row in rows:
        standards = _parse_standards(row["applicable_standards"])
        if standards:
            with_standards += 1
            for s in standards:
                by_standard_deliv[s].add(row["id"])
                if row["trade_value"]:
                    by_standard_trades[s].add(row["trade_value"])
                if row["service_value"]:
                    by_standard_services[s].add(row["service_value"])
        else:
            if len(without_standards_sample) < _SAMPLE_LIMIT:
                without_standards_sample.append(
                    {
                        "id": row["id"],
                        "summary": row["deliverables_summary"] or "",
                        "trade": row["trade_value"] or "",
                        "service": row["service_value"] or "",
                        "source": row["source_filename"] or "",
                    }
                )

    cells: list[ComplianceMatrixCell] = []
    for std, ids in by_standard_deliv.items():
        cells.append(
            ComplianceMatrixCell(
                standard=std,
                deliverable_count=len(ids),
                trades=sorted(by_standard_trades[std]),
                services=sorted(by_standard_services[std]),
            )
        )
    cells.sort(key=lambda c: (-c.deliverable_count, c.standard.lower()))

    one_off = sorted(
        [c.standard for c in cells if c.deliverable_count == 1],
        key=str.lower,
    )

    return ComplianceTraceabilityResult(
        project_name=project_name,
        generated_at=_now(),
        total_deliverables=total_deliverables,
        deliverables_with_standards=with_standards,
        deliverables_without_standards=total_deliverables - with_standards,
        distinct_standards=len(cells),
        standards=cells,
        deliverables_without_standards_sample=without_standards_sample,
        one_off_standards=one_off,
    )


# ─── Excel render ──────────────────────────────────────────────────────────


def _style_header(sheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22


def _write_matrix_sheet(
    workbook: Workbook, result: ComplianceTraceabilityResult
) -> None:
    sheet = workbook.active
    sheet.title = "Standards Matrix"
    headers = ["Standard", "# deliverables", "Trades touched", "Services touched"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    one_off = set(result.one_off_standards)
    for cell in result.standards:
        sheet.append(
            [
                cell.standard,
                cell.deliverable_count,
                ", ".join(cell.trades),
                ", ".join(cell.services),
            ]
        )
        if cell.standard in one_off:
            row_idx = sheet.max_row
            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = _ONE_OFF_FILL

    widths = [38, 16, 40, 40]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"


def _write_without_standards_sheet(
    workbook: Workbook, result: ComplianceTraceabilityResult
) -> None:
    sheet = workbook.create_sheet(title="Without Standards")
    headers = ["ID", "Trade", "Service", "Source", "Summary"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    for item in result.deliverables_without_standards_sample:
        sheet.append(
            [
                item.get("id", ""),
                item.get("trade", ""),
                item.get("service", ""),
                item.get("source", ""),
                item.get("summary", ""),
            ]
        )
    widths = [38, 22, 22, 38, 90]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"


def _write_summary_sheet(
    workbook: Workbook, result: ComplianceTraceabilityResult
) -> None:
    sheet = workbook.create_sheet(title="Summary")
    sheet.append(["Compliance Traceability — Summary"])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Project", result.project_name])
    sheet.append(["Generated at", result.generated_at])
    sheet.append([])
    sheet.append(["Total deliverables", result.total_deliverables])
    sheet.append(["Deliverables with standards", result.deliverables_with_standards])
    sheet.append(
        ["Deliverables without standards", result.deliverables_without_standards]
    )
    sheet.append(["Distinct standards cited", result.distinct_standards])
    sheet.append(["One-off standards (cited once)", len(result.one_off_standards)])
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 22


def write_xlsx(
    result: ComplianceTraceabilityResult, *, output_path: Path
) -> Path:
    """Write the compliance traceability workbook. Returns the path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _write_matrix_sheet(workbook, result)
    _write_without_standards_sheet(workbook, result)
    _write_summary_sheet(workbook, result)
    workbook.save(str(output_path))
    return output_path


__all__ = [
    "ComplianceMatrixCell",
    "ComplianceTraceabilityResult",
    "compute",
    "write_xlsx",
]
