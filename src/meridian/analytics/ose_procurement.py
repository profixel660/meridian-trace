"""OSE Procurement Completeness — checklist per equipment-vendor trade.

Per CONTEXT.md §5 vendor-attribution rule: equipment + native docs (shop
drawings, O&M, BIM, EPD, warranty, certification) → `<Class> Vendor`. There
is no generic "OSE Vendor"; each equipment class has its own vendor row in
the trade taxonomy (`category_hint = 'vendor'`).

For each vendor that has at least one deliverable in the project, we run a
fixed checklist of expected vendor-scope items and report which are present.
Matching is case-insensitive substring on `deliverables_summary` (and, for
the assembly check, the deliverable's category being `procurement`).

Pure read-only over the per-project SQLite. No LLM, no schema mutation.
"""

from __future__ import annotations

import sqlite3
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# Excel render style — kept consistent with src/meridian/export/excel.py.
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_AMBER_FILL = PatternFill("solid", fgColor="FFE0B2")  # rows below 60%

# Expected vendor-scope checklist + match rules (all defined here so they're
# easy to tune). Each entry: scope_item -> (kind, payload).
#   kind = 'category'  → match if deliverable.category = payload
#   kind = 'substring' → match if any string in payload is a case-insensitive
#                        substring of deliverables_summary
EXPECTED_VENDOR_SCOPE: list[tuple[str, str, tuple]] = [
    ("equipment_assembly", "category", ("procurement",)),
    ("shop_drawings", "substring", ("shop drawings", "technical datasheet")),
    (
        "operating_maintenance",
        "substring",
        ("o&m", "operating", "maintenance manual"),
    ),
    ("bim_model", "substring", ("bim", "lod300")),
    (
        "environmental_declaration",
        "substring",
        ("epd", "environmental product declaration"),
    ),
    ("warranty", "substring", ("warranty",)),
    (
        "functional_description",
        "substring",
        ("functional description", "fds"),
    ),
]

_PREVIEW_LEN = 120


# ─── Pydantic models ───────────────────────────────────────────────────────


class VendorScopeCheck(BaseModel):
    scope_item: str
    expected: bool = True
    found: bool
    matched_deliverable_id: str | None
    matched_summary_preview: str | None


class VendorCompleteness(BaseModel):
    vendor_trade: str
    deliverables_count: int
    scope_checklist: list[VendorScopeCheck]
    completeness_pct: float


class OseProcurementResult(BaseModel):
    project_name: str
    generated_at: str
    vendor_trades_present: list[str]
    vendors: list[VendorCompleteness]
    overall_completeness_pct: float


# ─── Internal helpers ──────────────────────────────────────────────────────


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _pct(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100.0, 1)


def _preview(text: str | None) -> str | None:
    if not text:
        return None
    text = text.strip()
    if len(text) <= _PREVIEW_LEN:
        return text
    return text[: _PREVIEW_LEN - 1] + "…"


def _match_one(
    deliverable: dict, kind: str, payload: tuple
) -> tuple[bool, str | None, str | None]:
    """Return (found, matched_id, matched_summary) for a scope item."""
    summary = (deliverable.get("summary") or "").lower()
    category = (deliverable.get("category") or "").lower()
    if kind == "category":
        wanted = payload[0].lower()
        if category == wanted:
            return True, deliverable["id"], deliverable.get("summary")
    elif kind == "substring":
        for needle in payload:
            if needle.lower() in summary:
                return True, deliverable["id"], deliverable.get("summary")
    return False, None, None


# ─── Public API ───────────────────────────────────────────────────────────


def compute(conn: sqlite3.Connection) -> OseProcurementResult:
    """Build the OSE procurement completeness pivot."""
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    project_row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    project_name = project_row["name"] if project_row else "(unknown)"

    # Pull every deliverable joined to its trade + category.
    rows = conn.execute(
        """
        SELECT
            d.id,
            d.deliverables_summary AS summary,
            tt.value AS trade_value,
            tt.category_hint AS trade_category_hint,
            ct.value AS category_value
        FROM deliverable d
        LEFT JOIN trade_taxonomy    tt ON tt.id = d.trade_id
        LEFT JOIN category_taxonomy ct ON ct.id = d.category_id
        WHERE tt.category_hint = 'vendor'
        """
    ).fetchall()

    # Group deliverables by vendor trade.
    by_vendor: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_vendor[row["trade_value"]].append(
            {
                "id": row["id"],
                "summary": row["summary"],
                "category": row["category_value"],
            }
        )

    vendor_trades_present = sorted(by_vendor.keys())

    vendors: list[VendorCompleteness] = []
    for vendor in vendor_trades_present:
        deliverables = by_vendor[vendor]
        checklist: list[VendorScopeCheck] = []
        for scope_item, kind, payload in EXPECTED_VENDOR_SCOPE:
            found = False
            matched_id: str | None = None
            matched_summary: str | None = None
            for deliv in deliverables:
                ok, m_id, m_summary = _match_one(deliv, kind, payload)
                if ok:
                    found = True
                    matched_id = m_id
                    matched_summary = m_summary
                    break
            checklist.append(
                VendorScopeCheck(
                    scope_item=scope_item,
                    expected=True,
                    found=found,
                    matched_deliverable_id=matched_id,
                    matched_summary_preview=_preview(matched_summary),
                )
            )
        found_count = sum(1 for c in checklist if c.found)
        vendors.append(
            VendorCompleteness(
                vendor_trade=vendor,
                deliverables_count=len(deliverables),
                scope_checklist=checklist,
                completeness_pct=_pct(found_count, len(checklist)),
            )
        )

    overall_pct = (
        round(sum(v.completeness_pct for v in vendors) / len(vendors), 1)
        if vendors
        else 0.0
    )

    return OseProcurementResult(
        project_name=project_name,
        generated_at=_now(),
        vendor_trades_present=vendor_trades_present,
        vendors=vendors,
        overall_completeness_pct=overall_pct,
    )


# ─── Excel render ──────────────────────────────────────────────────────────


def _style_header(sheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22


def _write_completeness_sheet(
    workbook: Workbook, result: OseProcurementResult
) -> None:
    sheet = workbook.active
    sheet.title = "Vendor Completeness"

    scope_items = [item for item, _kind, _payload in EXPECTED_VENDOR_SCOPE]
    headers = ["Vendor Trade", *scope_items, "Completeness %"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    for vendor in result.vendors:
        check_by_item = {c.scope_item: c for c in vendor.scope_checklist}
        row_values = [vendor.vendor_trade]
        for item in scope_items:
            check = check_by_item.get(item)
            row_values.append("✓" if check and check.found else "")
        row_values.append(vendor.completeness_pct)
        sheet.append(row_values)
        if vendor.completeness_pct < 60.0:
            row_idx = sheet.max_row
            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = _AMBER_FILL

    sheet.column_dimensions["A"].width = 26
    for i in range(2, len(headers)):
        sheet.column_dimensions[get_column_letter(i)].width = 22
    sheet.column_dimensions[get_column_letter(len(headers))].width = 16
    sheet.freeze_panes = "B2"


def _write_detail_sheet(
    workbook: Workbook, result: OseProcurementResult
) -> None:
    sheet = workbook.create_sheet(title="Vendor Detail")
    headers = ["Vendor Trade", "Scope item", "Found", "Deliverable ID", "Summary preview"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    for i, vendor in enumerate(result.vendors):
        if i > 0:
            sheet.append([])  # blank row between vendors
        for check in vendor.scope_checklist:
            sheet.append(
                [
                    vendor.vendor_trade,
                    check.scope_item,
                    "✓" if check.found else "",
                    check.matched_deliverable_id or "",
                    check.matched_summary_preview or "",
                ]
            )

    widths = [26, 26, 8, 38, 90]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"


def _write_summary_sheet(
    workbook: Workbook, result: OseProcurementResult
) -> None:
    sheet = workbook.create_sheet(title="Summary")
    sheet.append(["OSE Procurement Completeness — Summary"])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Project", result.project_name])
    sheet.append(["Generated at", result.generated_at])
    sheet.append([])
    sheet.append(["Vendors present", len(result.vendors)])
    sheet.append(
        [
            "Vendors at 100%",
            sum(1 for v in result.vendors if v.completeness_pct >= 100.0),
        ]
    )
    sheet.append(
        [
            "Vendors below 60%",
            sum(1 for v in result.vendors if v.completeness_pct < 60.0),
        ]
    )
    sheet.append(["Overall completeness %", result.overall_completeness_pct])
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 22


def write_xlsx(result: OseProcurementResult, *, output_path: Path) -> Path:
    """Write the OSE procurement completeness workbook. Returns the path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _write_completeness_sheet(workbook, result)
    _write_detail_sheet(workbook, result)
    _write_summary_sheet(workbook, result)
    workbook.save(str(output_path))
    return output_path


__all__ = [
    "EXPECTED_VENDOR_SCOPE",
    "OseProcurementResult",
    "VendorCompleteness",
    "VendorScopeCheck",
    "compute",
    "write_xlsx",
]
