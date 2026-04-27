"""Trade Overlap — surface where two trades touch the same scope item.

Driven by `deliverable.extraction_group_id`: rows that share an extraction
group came from the same candidate that fanned out across multiple trades
(per CONTEXT.md §4 multi-row rule). Two trades that co-occur on the same
extraction_group_id are by construction touching the same scope item.

Pure read-only over the per-project SQLite. No LLM, no schema mutation.
"""

from __future__ import annotations

import sqlite3
from collections import defaultdict
from datetime import UTC, datetime
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# Excel render style — kept consistent with src/meridian/export/excel.py.
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Heat-fill palette for co-occurrence counts (light → dark amber).
_HEAT_PALETTE = ["FFF7E0", "FFE9AD", "FFD27A", "FFB347", "FF8C42", "E46A1E"]

_SAMPLE_PER_PAIR = 5
_NULL_TRADE_LABEL = "(null)"


# ─── Pydantic models ───────────────────────────────────────────────────────


class TradeOverlapPair(BaseModel):
    trade_a: str
    trade_b: str
    co_occurrence_count: int
    sample_summaries: list[str]


class TradeOverlapResult(BaseModel):
    project_name: str
    generated_at: str
    total_extraction_groups: int
    multi_trade_groups: int
    multi_trade_pct: float
    pairs: list[TradeOverlapPair]
    trade_distribution: list[dict]


# ─── Internal helpers ──────────────────────────────────────────────────────


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _pct(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100.0, 1)


# ─── Public API ───────────────────────────────────────────────────────────


def compute(conn: sqlite3.Connection) -> TradeOverlapResult:
    """Compute trade-overlap pairs and per-trade distribution."""
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    project_row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    project_name = project_row["name"] if project_row else "(unknown)"

    rows = conn.execute(
        """
        SELECT
            d.extraction_group_id,
            d.deliverables_summary,
            tt.value AS trade_value
        FROM deliverable d
        LEFT JOIN trade_taxonomy tt ON tt.id = d.trade_id
        """
    ).fetchall()

    # group_id -> set of trade values (None preserved separately)
    group_trades: dict[str, set[str]] = defaultdict(set)
    group_has_null: dict[str, bool] = defaultdict(bool)
    group_summaries: dict[str, list[str]] = defaultdict(list)

    for row in rows:
        gid = row["extraction_group_id"]
        if gid is None:
            continue
        trade = row["trade_value"]
        if trade is None:
            group_has_null[gid] = True
        else:
            group_trades[gid].add(trade)
        if row["deliverables_summary"]:
            group_summaries[gid].append(row["deliverables_summary"])

    all_group_ids = set(group_trades.keys()) | set(group_has_null.keys())
    total_groups = len(all_group_ids)
    multi_trade_groups = sum(1 for gid in all_group_ids if len(group_trades[gid]) > 1)

    # Pair detection — ignore NULL-trade rows (per spec). Sort tuple alpha
    # before counting to keep the matrix symmetric with one row per pair.
    pair_count: dict[tuple[str, str], int] = defaultdict(int)
    pair_samples: dict[tuple[str, str], list[str]] = defaultdict(list)

    for gid, trades in group_trades.items():
        if len(trades) < 2:
            continue
        sorted_trades = sorted(trades)
        for a, b in combinations(sorted_trades, 2):
            key = (a, b)
            pair_count[key] += 1
            samples = pair_samples[key]
            if len(samples) < _SAMPLE_PER_PAIR and group_summaries.get(gid):
                # Use the first non-empty summary from this group as a sample.
                samples.append(group_summaries[gid][0])

    pairs = [
        TradeOverlapPair(
            trade_a=a,
            trade_b=b,
            co_occurrence_count=count,
            sample_summaries=pair_samples[(a, b)][:_SAMPLE_PER_PAIR],
        )
        for (a, b), count in pair_count.items()
    ]
    pairs.sort(key=lambda p: (-p.co_occurrence_count, p.trade_a, p.trade_b))

    # Trade distribution — count of distinct extraction_group_ids each trade
    # appears in. NULL-trade is included as `(null)` for visibility.
    trade_to_groups: dict[str, set[str]] = defaultdict(set)
    for gid, trades in group_trades.items():
        for t in trades:
            trade_to_groups[t].add(gid)
    for gid, has_null in group_has_null.items():
        if has_null:
            trade_to_groups[_NULL_TRADE_LABEL].add(gid)

    trade_distribution = [
        {
            "trade": trade,
            "group_count": len(groups),
            "pct_of_groups": _pct(len(groups), total_groups),
        }
        for trade, groups in trade_to_groups.items()
    ]
    trade_distribution.sort(key=lambda d: (-d["group_count"], str(d["trade"]).lower()))

    return TradeOverlapResult(
        project_name=project_name,
        generated_at=_now(),
        total_extraction_groups=total_groups,
        multi_trade_groups=multi_trade_groups,
        multi_trade_pct=_pct(multi_trade_groups, total_groups),
        pairs=pairs,
        trade_distribution=trade_distribution,
    )


# ─── Excel render ──────────────────────────────────────────────────────────


def _style_header(sheet, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22


def _heat_fill_for(value: int, max_value: int) -> PatternFill | None:
    if max_value <= 0 or value <= 0:
        return None
    # Map value into one of the palette buckets.
    idx = min(
        len(_HEAT_PALETTE) - 1,
        int((value / max_value) * len(_HEAT_PALETTE)),
    )
    return PatternFill("solid", fgColor=_HEAT_PALETTE[idx])


def _write_pairs_sheet(workbook: Workbook, result: TradeOverlapResult) -> None:
    sheet = workbook.active
    sheet.title = "Trade Overlap Pairs"
    headers = ["Trade A", "Trade B", "Co-occurrence count", "Sample summaries"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    max_count = max((p.co_occurrence_count for p in result.pairs), default=0)
    for pair in result.pairs:
        sheet.append(
            [
                pair.trade_a,
                pair.trade_b,
                pair.co_occurrence_count,
                "; ".join(pair.sample_summaries),
            ]
        )
        fill = _heat_fill_for(pair.co_occurrence_count, max_count)
        if fill is not None:
            sheet.cell(row=sheet.max_row, column=3).fill = fill

    widths = [26, 26, 22, 100]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"


def _write_distribution_sheet(
    workbook: Workbook, result: TradeOverlapResult
) -> None:
    sheet = workbook.create_sheet(title="Trade Distribution")
    headers = ["Trade", "# distinct scope groups", "% of all groups"]
    sheet.append(headers)
    _style_header(sheet, len(headers))

    for entry in result.trade_distribution:
        sheet.append(
            [
                entry["trade"],
                entry["group_count"],
                entry["pct_of_groups"],
            ]
        )

    widths = [28, 26, 18]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.freeze_panes = "A2"


def _write_summary_sheet(
    workbook: Workbook, result: TradeOverlapResult
) -> None:
    sheet = workbook.create_sheet(title="Summary")
    sheet.append(["Trade Overlap — Summary"])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.append([])
    sheet.append(["Project", result.project_name])
    sheet.append(["Generated at", result.generated_at])
    sheet.append([])
    sheet.append(["Total extraction groups", result.total_extraction_groups])
    sheet.append(["Multi-trade groups", result.multi_trade_groups])
    sheet.append(["Multi-trade %", result.multi_trade_pct])
    sheet.append([])
    sheet.append(["Top 3 overlap pairs:"])
    for pair in result.pairs[:3]:
        sheet.append(
            [
                f"{pair.trade_a} / {pair.trade_b}",
                pair.co_occurrence_count,
            ]
        )
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 22


def write_xlsx(result: TradeOverlapResult, *, output_path: Path) -> Path:
    """Write the trade-overlap workbook. Returns the path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _write_pairs_sheet(workbook, result)
    _write_distribution_sheet(workbook, result)
    _write_summary_sheet(workbook, result)
    workbook.save(str(output_path))
    return output_path


__all__ = [
    "TradeOverlapPair",
    "TradeOverlapResult",
    "compute",
    "write_xlsx",
]
