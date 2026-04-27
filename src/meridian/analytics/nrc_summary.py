"""NRC Scope-Shift Summary — every deliverable carrying `scope_shifted_to_nrc`.

Useful for tenant fit-out conversations and base-build vs NRC scope
clarification (CONTEXT.md §8 — `scope_shifted_to_nrc` flag).

Pure read-only SQL aggregation. Considers EVERY deliverable row regardless
of status — an NRC-flagged item that is still quarantined is exactly the
sort of thing this report exists to surface.

The summary preserves the ⚠ marker convention for `negotiated_response`
deliverables (CONTEXT.md §4) — applied at render time here, since the
flag almost always co-occurs with `scope_shifted_to_nrc`.
"""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# ─── Models ────────────────────────────────────────────────────────────────


class NrcScopeItem(BaseModel):
    deliverable_id: str
    summary: str
    trade: str | None
    service: str | None
    source_filename: str
    source_ref: str
    nrc_context: str


class NrcSummaryResult(BaseModel):
    project_name: str
    generated_at: str
    total_items: int
    items_by_trade: dict[str, list[NrcScopeItem]]
    all_items: list[NrcScopeItem]


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_NRC_FLAG = "scope_shifted_to_nrc"
_NEGOTIATED_FLAG = "negotiated_response"
_WARN_MARKER = " \u26a0"  # ⚠


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _project_name(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    return row["name"] if row else "(unknown)"


def _safe_loads(raw: str | None, default):
    if not raw:
        return default
    try:
        out = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return default
    return out


def _render_source_ref(raw: str | None) -> str:
    obj = _safe_loads(raw, {})
    if isinstance(obj, dict):
        return str(obj.get("rendered") or "")
    return ""


def _nrc_context_text(flag_context_obj: dict) -> str:
    """Pull the verbatim NRC context payload out of flag_context.

    flag_context[scope_shifted_to_nrc] is typically a string but may be a
    dict for richer payloads — render either to a flat readable string.
    """
    raw = flag_context_obj.get(_NRC_FLAG)
    if raw is None:
        return ""
    if isinstance(raw, str):
        return raw
    if isinstance(raw, dict):
        # Prefer a 'text'/'description' field; fall back to compact JSON.
        for k in ("text", "description", "summary", "value"):
            if k in raw and isinstance(raw[k], str):
                return raw[k]
        return json.dumps(raw, ensure_ascii=False)
    return str(raw)


# ─── Public API ────────────────────────────────────────────────────────────


def compute(conn: sqlite3.Connection) -> NrcSummaryResult:
    """Return every deliverable row carrying `scope_shifted_to_nrc`."""
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    rows = conn.execute(
        """
        SELECT
            d.id                    AS deliverable_id,
            d.deliverables_summary  AS summary,
            d.flags                 AS flags,
            d.flag_context          AS flag_context,
            d.source_ref            AS source_ref,
            sd.filename             AS source_filename,
            tt.value                AS trade,
            st.value                AS service
        FROM deliverable d
        LEFT JOIN source_document   sd ON sd.id = d.source_id
        LEFT JOIN trade_taxonomy    tt ON tt.id = d.trade_id
        LEFT JOIN service_taxonomy  st ON st.id = d.service_id
        WHERE d.flags LIKE '%' || ? || '%'
        ORDER BY tt.value, sd.filename, d.created_at
        """,
        (_NRC_FLAG,),
    ).fetchall()

    items: list[NrcScopeItem] = []
    by_trade: dict[str, list[NrcScopeItem]] = defaultdict(list)
    for r in rows:
        flags = _safe_loads(r["flags"], [])
        if not isinstance(flags, list) or _NRC_FLAG not in flags:
            # LIKE match was approximate — confirm against parsed flags array.
            continue
        flag_ctx = _safe_loads(r["flag_context"], {})
        if not isinstance(flag_ctx, dict):
            flag_ctx = {}

        summary = r["summary"] or ""
        if _NEGOTIATED_FLAG in flags and _WARN_MARKER.strip() not in summary:
            summary = f"{summary}{_WARN_MARKER}"

        item = NrcScopeItem(
            deliverable_id=r["deliverable_id"],
            summary=summary,
            trade=r["trade"],
            service=r["service"],
            source_filename=r["source_filename"] or "",
            source_ref=_render_source_ref(r["source_ref"]),
            nrc_context=_nrc_context_text(flag_ctx),
        )
        items.append(item)
        by_trade[item.trade or "(unset)"].append(item)

    return NrcSummaryResult(
        project_name=_project_name(conn),
        generated_at=_now(),
        total_items=len(items),
        items_by_trade=dict(by_trade),
        all_items=items,
    )


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22


def write_xlsx(result: NrcSummaryResult, *, output_path: Path) -> Path:
    """Render two sheets: 'NRC Scope Items' (flat) + 'By Trade' (pivot)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    flat = wb.active
    flat.title = "NRC Scope Items"
    headers = ["Trade", "Service", "Source", "Summary", "NRC context"]
    flat.append(headers)
    _style_header(flat)
    for item in result.all_items:
        flat.append(
            [
                item.trade or "",
                item.service or "",
                item.source_filename
                + (f" ({item.source_ref})" if item.source_ref else ""),
                item.summary,
                item.nrc_context,
            ]
        )
    widths = {
        "Trade": 22,
        "Service": 22,
        "Source": 50,
        "Summary": 70,
        "NRC context": 70,
    }
    for col_index, name in enumerate(headers, start=1):
        flat.column_dimensions[get_column_letter(col_index)].width = widths[name]
    flat.freeze_panes = "A2"

    pivot = wb.create_sheet("By Trade")
    pivot.append(["Trade", "Count", "Summary preview"])
    _style_header(pivot)
    for trade in sorted(result.items_by_trade):
        bucket = result.items_by_trade[trade]
        preview = "; ".join(it.summary for it in bucket[:3])
        pivot.append([trade, len(bucket), preview])
    pivot.column_dimensions["A"].width = 28
    pivot.column_dimensions["B"].width = 8
    pivot.column_dimensions["C"].width = 100
    pivot.freeze_panes = "A2"

    wb.save(str(output_path))
    return output_path


__all__ = [
    "NrcScopeItem",
    "NrcSummaryResult",
    "compute",
    "write_xlsx",
]
