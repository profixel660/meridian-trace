"""Conflict Register Export — one-stop conflict log for QA / claims teams.

Pure read-only SQL aggregation against the `conflict` and `conflict_party`
tables (DATA_MODEL_V1.md §9). Renders a three-sheet xlsx:

    Conflicts  — one row per conflict, status + most-onerous reasoning.
    Parties    — one row per party (so each conflict appears 2+ times).
    Summary    — counts by kind / status + a one-sentence headline.

Mirrors the party-resolution pattern from `review/conflicts.py::_load_parties`:
deliverable parties are joined to deliverable / source / trade / service;
audit parties are joined to audit_record / source.
"""

from __future__ import annotations

import sqlite3
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# ─── Models ────────────────────────────────────────────────────────────────


class ConflictRegisterParty(BaseModel):
    party_kind: Literal["deliverable", "audit"]
    party_id: str
    position: str | None
    summary_or_text: str
    source_filename: str | None
    source_ref: str | None
    trade: str | None
    service: str | None


class ConflictRegisterEntry(BaseModel):
    conflict_id: str
    kind: Literal[
        "cross_source_content",
        "responsibility",
        "revision",
        "document_class",
        "scope_demarcation",
    ]
    status: str
    most_onerous_party_id: str | None
    most_onerous_reasoning: str
    is_directly_comparable: bool
    parties: list[ConflictRegisterParty]
    created_at: str
    resolved_at: str | None


class ConflictRegisterResult(BaseModel):
    project_name: str
    generated_at: str
    total: int
    pending: int
    by_kind: dict[str, int]
    entries: list[ConflictRegisterEntry]


# ─── Style ─────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PENDING_FILL = PatternFill("solid", fgColor="FFE9B0")


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _project_name(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    return row["name"] if row else "(unknown)"


def _render_source_ref(raw: str | None) -> str | None:
    if not raw:
        return None
    import json as _json

    try:
        obj = _json.loads(raw)
    except (ValueError, TypeError):
        return None
    if isinstance(obj, dict):
        rendered = obj.get("rendered")
        return str(rendered) if rendered else None
    return None


def _load_parties(
    conn: sqlite3.Connection, conflict_id: str
) -> list[ConflictRegisterParty]:
    party_rows = conn.execute(
        "SELECT party_kind, party_id, party_position FROM conflict_party "
        "WHERE conflict_id = ? ORDER BY rowid",
        (conflict_id,),
    ).fetchall()
    out: list[ConflictRegisterParty] = []
    for pr in party_rows:
        kind = pr["party_kind"]
        if kind == "deliverable":
            d = conn.execute(
                """
                SELECT d.deliverables_summary, d.source_ref,
                       sd.filename AS source_filename,
                       tt.value    AS trade_value,
                       st.value    AS service_value
                FROM deliverable d
                LEFT JOIN source_document  sd ON sd.id = d.source_id
                LEFT JOIN trade_taxonomy   tt ON tt.id = d.trade_id
                LEFT JOIN service_taxonomy st ON st.id = d.service_id
                WHERE d.id = ?
                """,
                (pr["party_id"],),
            ).fetchone()
            if d is not None:
                out.append(
                    ConflictRegisterParty(
                        party_kind="deliverable",
                        party_id=pr["party_id"],
                        position=pr["party_position"],
                        summary_or_text=d["deliverables_summary"] or "",
                        source_filename=d["source_filename"],
                        source_ref=_render_source_ref(d["source_ref"]),
                        trade=d["trade_value"],
                        service=d["service_value"],
                    )
                )
                continue
            # Deliverable row missing — fall through to a stub so the
            # conflict is still rendered.
            out.append(
                ConflictRegisterParty(
                    party_kind="deliverable",
                    party_id=pr["party_id"],
                    position=pr["party_position"],
                    summary_or_text="(deliverable row not found)",
                    source_filename=None,
                    source_ref=None,
                    trade=None,
                    service=None,
                )
            )
        elif kind == "audit":
            a = conn.execute(
                """
                SELECT a.candidate_text, a.source_ref,
                       sd.filename AS source_filename
                FROM audit_record a
                LEFT JOIN source_document sd ON sd.id = a.source_id
                WHERE a.id = ?
                """,
                (pr["party_id"],),
            ).fetchone()
            text = a["candidate_text"] if a is not None else "(audit row not found)"
            filename = a["source_filename"] if a is not None else None
            ref_raw = a["source_ref"] if a is not None else None
            out.append(
                ConflictRegisterParty(
                    party_kind="audit",
                    party_id=pr["party_id"],
                    position=pr["party_position"],
                    summary_or_text=text or "",
                    source_filename=filename,
                    source_ref=_render_source_ref(ref_raw),
                    trade=None,
                    service=None,
                )
            )
        else:
            # Unknown party_kind — keep it visible in the export.
            out.append(
                ConflictRegisterParty(
                    party_kind=kind,  # type: ignore[arg-type]
                    party_id=pr["party_id"],
                    position=pr["party_position"],
                    summary_or_text="",
                    source_filename=None,
                    source_ref=None,
                    trade=None,
                    service=None,
                )
            )
    return out


# ─── Public API ────────────────────────────────────────────────────────────


def compute(
    conn: sqlite3.Connection, *, status_filter: str | None = None
) -> ConflictRegisterResult:
    """Build the conflict register.

    Sort order: pending first by created_at ASC, then resolved by
    resolved_at DESC (most recently resolved first). NULL resolved_at on a
    "resolved" status (shouldn't happen but guard anyway) sinks to the end.
    """
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    sql = "SELECT * FROM conflict"
    params: tuple = ()
    if status_filter is not None:
        sql += " WHERE status = ?"
        params = (status_filter,)

    rows = conn.execute(sql, params).fetchall()

    entries: list[ConflictRegisterEntry] = []
    by_kind: Counter[str] = Counter()
    pending = 0

    for r in rows:
        parties = _load_parties(conn, r["id"])
        most_onerous = r["most_onerous_party_id"]
        entries.append(
            ConflictRegisterEntry(
                conflict_id=r["id"],
                kind=r["kind"],
                status=r["status"],
                most_onerous_party_id=most_onerous,
                most_onerous_reasoning=r["most_onerous_reasoning"] or "",
                is_directly_comparable=most_onerous is not None,
                parties=parties,
                created_at=r["created_at"],
                resolved_at=r["resolved_at"],
            )
        )
        by_kind[r["kind"]] += 1
        if r["status"] == "pending":
            pending += 1

    def _sort_key(e: ConflictRegisterEntry) -> tuple:
        # pending (0) before resolved (1); within pending sort by created_at ASC;
        # within resolved sort by resolved_at DESC (negate via reverse string).
        if e.status == "pending":
            return (0, e.created_at, e.conflict_id)
        # For DESC by resolved_at among resolved: invert by using negative
        # ordering — easier: sort tuple uses a flipped resolver.
        # Trick: return (1, -timestamp_as_int) — but timestamps are strings.
        # Use the negated string sort by chr-inverting; cleaner: sort resolved
        # in a separate pass below.
        return (1, e.resolved_at or "", e.conflict_id)

    pending_entries = [e for e in entries if e.status == "pending"]
    resolved_entries = [e for e in entries if e.status != "pending"]
    pending_entries.sort(key=lambda e: (e.created_at, e.conflict_id))
    resolved_entries.sort(
        key=lambda e: (e.resolved_at or "", e.conflict_id), reverse=True
    )
    ordered = pending_entries + resolved_entries

    return ConflictRegisterResult(
        project_name=_project_name(conn),
        generated_at=_now(),
        total=len(entries),
        pending=pending,
        by_kind=dict(by_kind),
        entries=ordered,
    )


# ─── Excel renderer ────────────────────────────────────────────────────────


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22


def _write_summary_sheet(wb: Workbook, result: ConflictRegisterResult) -> None:
    sheet = wb.create_sheet("Summary", 0)
    sheet.append(["Conflict Register — Summary"])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([f"Project: {result.project_name}"])
    sheet.append([f"Generated: {result.generated_at}"])
    sheet.append([])

    not_comparable = sum(
        1 for e in result.entries if not e.is_directly_comparable
    )
    resolved = result.total - result.pending
    headline = (
        f"{result.total} conflicts surfaced; {resolved} resolved; "
        f"{result.pending} pending; {not_comparable} classified as "
        "'not directly comparable'."
    )
    sheet.append([headline])
    sheet["A5"].font = Font(italic=True)
    sheet.append([])

    sheet.append(["By kind"])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
    for kind in sorted(result.by_kind):
        sheet.append([kind, result.by_kind[kind]])

    sheet.append([])
    sheet.append(["By status"])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
    status_counts: Counter[str] = Counter(e.status for e in result.entries)
    for status in sorted(status_counts):
        sheet.append([status, status_counts[status]])

    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 14


def _write_conflicts_sheet(wb: Workbook, result: ConflictRegisterResult) -> None:
    sheet = wb.create_sheet("Conflicts")
    headers = [
        "Status",
        "Kind",
        "Most onerous?",
        "Reasoning",
        "Resolved at",
        "Conflict ID",
    ]
    sheet.append(headers)
    _style_header(sheet)
    for e in result.entries:
        most_onerous_label = "yes" if e.is_directly_comparable else "not directly comparable"
        sheet.append(
            [
                e.status,
                e.kind,
                most_onerous_label,
                e.most_onerous_reasoning,
                e.resolved_at or "",
                e.conflict_id,
            ]
        )
        if e.status == "pending":
            for col in range(1, len(headers) + 1):
                sheet.cell(row=sheet.max_row, column=col).fill = _PENDING_FILL
    widths = {
        "Status": 14,
        "Kind": 22,
        "Most onerous?": 24,
        "Reasoning": 80,
        "Resolved at": 22,
        "Conflict ID": 38,
    }
    for col_index, name in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]
    sheet.freeze_panes = "A2"


def _write_parties_sheet(wb: Workbook, result: ConflictRegisterResult) -> None:
    sheet = wb.create_sheet("Parties")
    headers = [
        "Conflict ID",
        "Party",
        "Trade/Service",
        "Source",
        "Position",
        "Summary",
    ]
    sheet.append(headers)
    _style_header(sheet)
    for e in result.entries:
        for p in e.parties:
            party_label = p.party_kind
            trade_service = ""
            if p.party_kind == "deliverable":
                t = p.trade or ""
                s = p.service or ""
                trade_service = " / ".join(x for x in (t, s) if x)
            source = p.source_filename or ""
            if p.source_ref:
                source = f"{source} ({p.source_ref})" if source else p.source_ref
            sheet.append(
                [
                    e.conflict_id,
                    party_label,
                    trade_service,
                    source,
                    p.position or "",
                    p.summary_or_text,
                ]
            )
    widths = {
        "Conflict ID": 38,
        "Party": 12,
        "Trade/Service": 28,
        "Source": 50,
        "Position": 60,
        "Summary": 80,
    }
    for col_index, name in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]
    sheet.freeze_panes = "A2"


def write_xlsx(result: ConflictRegisterResult, *, output_path: Path) -> Path:
    """Write a three-sheet xlsx (Summary / Conflicts / Parties)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Drop the default empty sheet — we add our own in the order we want.
    default = wb.active
    wb.remove(default)

    _write_summary_sheet(wb, result)
    _write_conflicts_sheet(wb, result)
    _write_parties_sheet(wb, result)

    wb.save(str(output_path))
    return output_path


__all__ = [
    "ConflictRegisterEntry",
    "ConflictRegisterParty",
    "ConflictRegisterResult",
    "compute",
    "write_xlsx",
]
