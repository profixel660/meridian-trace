"""Risk Hotspot Map — top-N most-at-risk deliverables for PM prioritisation.

Pure read-only SQL aggregation. Risk score is a simple weighted sum of
quality / process signals already on each deliverable row:

    confidence == 'low'          → +3
    confidence == 'medium'       → +1
    each flag in flags array     → +1
    gate_outcome == 'borderline' → +2
    pending conflict participation → +5
    status == 'quarantined'      → +1

A deliverable participates in a pending conflict when its `flags` array
contains a `conflicts_with_source_<conflict_id>` token AND the matching
conflict row is still in `status = 'pending'`. We resolve that here against
the conflict table rather than trusting the flag in isolation, because
`review/conflicts.py::resolve_conflict` strips the flag at resolution time
— but in case any get out of sync, the conflict-table check is the source
of truth.
"""

from __future__ import annotations

import json
import sqlite3
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# ─── Models ────────────────────────────────────────────────────────────────


class RiskHotspot(BaseModel):
    deliverable_id: str
    summary: str
    trade: str | None
    service: str | None
    source_filename: str
    confidence: str
    flags: list[str]
    gate_outcome: str
    status: str
    risk_score: int
    risk_factors: list[str]


class RiskHotspotsResult(BaseModel):
    project_name: str
    generated_at: str
    items: list[RiskHotspot]
    score_distribution: dict[int, int]


# ─── Style (matches export/excel.py) ───────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_AMBER_FILL = PatternFill("solid", fgColor="FFE9B0")

_CONFIDENCE_WEIGHT = {"low": 3, "medium": 1, "high": 0}
_BORDERLINE_WEIGHT = 2
_PENDING_CONFLICT_WEIGHT = 5
_QUARANTINED_WEIGHT = 1
_CONFLICT_FLAG_PREFIX = "conflicts_with_source_"


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _safe_loads_list(raw: str | None) -> list:
    if not raw:
        return []
    try:
        out = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []
    return out if isinstance(out, list) else []


def _project_name(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    return row["name"] if row else "(unknown)"


def _pending_conflict_ids(conn: sqlite3.Connection) -> set[str]:
    rows = conn.execute("SELECT id FROM conflict WHERE status = 'pending'").fetchall()
    return {r["id"] for r in rows}


def _score_one(
    *,
    confidence: str,
    flags: list[str],
    gate_outcome: str,
    status: str,
    pending_conflicts: set[str],
) -> tuple[int, list[str]]:
    score = 0
    factors: list[str] = []
    cw = _CONFIDENCE_WEIGHT.get(confidence, 0)
    if cw:
        score += cw
        factors.append(f"confidence={confidence} (+{cw})")
    if flags:
        score += len(flags)
        factors.append(f"{len(flags)} flag(s) (+{len(flags)})")
    if gate_outcome == "borderline":
        score += _BORDERLINE_WEIGHT
        factors.append(f"borderline gate (+{_BORDERLINE_WEIGHT})")
    in_pending = False
    for f in flags:
        if f.startswith(_CONFLICT_FLAG_PREFIX):
            cid = f[len(_CONFLICT_FLAG_PREFIX) :]
            if cid in pending_conflicts:
                in_pending = True
                break
    if in_pending:
        score += _PENDING_CONFLICT_WEIGHT
        factors.append(f"pending conflict (+{_PENDING_CONFLICT_WEIGHT})")
    if status == "quarantined":
        score += _QUARANTINED_WEIGHT
        factors.append(f"quarantined (+{_QUARANTINED_WEIGHT})")
    return score, factors


# ─── Public API ────────────────────────────────────────────────────────────


def compute(conn: sqlite3.Connection, *, top_n: int = 50) -> RiskHotspotsResult:
    """Compute the top-N most-at-risk deliverables.

    Pure read-only. Considers EVERY row in `deliverable` regardless of status —
    quarantined and rejected rows are themselves risk signals, so they belong
    in the histogram even if they sit outside the master register.
    """
    if conn.row_factory is not sqlite3.Row:
        conn.row_factory = sqlite3.Row

    pending_conflicts = _pending_conflict_ids(conn)

    rows = conn.execute(
        """
        SELECT
            d.id                        AS deliverable_id,
            d.deliverables_summary      AS summary,
            d.confidence                AS confidence,
            d.flags                     AS flags,
            d.gate_outcome              AS gate_outcome,
            d.status                    AS status,
            sd.filename                 AS source_filename,
            tt.value                    AS trade,
            st.value                    AS service
        FROM deliverable d
        LEFT JOIN source_document   sd ON sd.id = d.source_id
        LEFT JOIN trade_taxonomy    tt ON tt.id = d.trade_id
        LEFT JOIN service_taxonomy  st ON st.id = d.service_id
        """
    ).fetchall()

    scored: list[RiskHotspot] = []
    histogram: Counter[int] = Counter()
    for r in rows:
        flags = [str(f) for f in _safe_loads_list(r["flags"])]
        score, factors = _score_one(
            confidence=r["confidence"] or "",
            flags=flags,
            gate_outcome=r["gate_outcome"] or "",
            status=r["status"] or "",
            pending_conflicts=pending_conflicts,
        )
        histogram[score] += 1
        scored.append(
            RiskHotspot(
                deliverable_id=r["deliverable_id"],
                summary=r["summary"] or "",
                trade=r["trade"],
                service=r["service"],
                source_filename=r["source_filename"] or "",
                confidence=r["confidence"] or "",
                flags=flags,
                gate_outcome=r["gate_outcome"] or "",
                status=r["status"] or "",
                risk_score=score,
                risk_factors=factors,
            )
        )

    scored.sort(key=lambda x: (-x.risk_score, x.deliverable_id))
    top = scored[:top_n]

    return RiskHotspotsResult(
        project_name=_project_name(conn),
        generated_at=_now(),
        items=top,
        score_distribution=dict(histogram),
    )


def write_xlsx(result: RiskHotspotsResult, *, output_path: Path) -> Path:
    """Render a single-sheet 'Risk Hotspots' xlsx. Returns the path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Risk Hotspots"

    headers = [
        "Rank",
        "Risk score",
        "Trade",
        "Service",
        "Confidence",
        "Flags",
        "Status",
        "Source",
        "Risk factors",
        "Summary",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    for rank, item in enumerate(result.items, start=1):
        sheet.append(
            [
                rank,
                item.risk_score,
                item.trade or "",
                item.service or "",
                item.confidence,
                ", ".join(item.flags),
                item.status,
                item.source_filename,
                "; ".join(item.risk_factors),
                item.summary,
            ]
        )
        if rank <= 10:
            sheet.cell(row=rank + 1, column=2).fill = _AMBER_FILL

    widths = {
        "Rank": 6,
        "Risk score": 11,
        "Trade": 20,
        "Service": 22,
        "Confidence": 11,
        "Flags": 32,
        "Status": 14,
        "Source": 38,
        "Risk factors": 40,
        "Summary": 80,
    }
    for col_index, name in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]
    sheet.freeze_panes = "A2"

    wb.save(str(output_path))
    return output_path


__all__ = [
    "RiskHotspot",
    "RiskHotspotsResult",
    "compute",
    "write_xlsx",
]
