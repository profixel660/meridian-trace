"""Tender package generation — per-trade extracts of the master register.

This module reads from the project SQLite (read-only) and writes a
tender-ready file in either xlsx or md format. The xlsx output mirrors the
styling of ``meridian.export.excel`` so a head contractor can paste sheets
straight into a tender pack.

Key conventions
---------------
- "Accepted" means the row is in ``v_master_register`` (statuses
  ``auto_approved``, ``user_accepted``, ``user_edited``, ``user_promoted``,
  with superseded parents filtered out). Quarantined and rejected rows
  never appear in a tender package.
- Output goes under ``<projects_dir>/<slug>.tenders/`` by convention,
  parallel to the existing ``<slug>.logs/`` layout. Callers may override
  with ``output_dir``.
- Filenames are ``<slug>__tender__<trade-slug>__<UTC-timestamp>.<ext>`` so
  successive runs do not silently overwrite earlier packages.

This file does not import the CLI / API surfaces — they import from here.
"""

from __future__ import annotations

import json
import re
import sqlite3
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from meridian.config import settings
from meridian.db.connection import connect
from meridian.logging import get_logger
from meridian.projects import project_db_path

_log = get_logger("meridian.tender")

# --------------------------------------------------------------------------
# Public constants
# --------------------------------------------------------------------------

#: Directory suffix appended to ``<projects_dir>/<slug>`` for tender outputs.
DEFAULT_TENDERS_SUBDIR = ".tenders"

#: Output formats supported by :func:`build_tender_package`.
SUPPORTED_FORMATS: tuple[str, ...] = ("xlsx", "md")

OutputFormat = Literal["xlsx", "md"]

# --------------------------------------------------------------------------
# Styling — kept aligned with meridian.export.excel for visual consistency
# --------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
_SUBHEADER_FONT = Font(bold=True, color="FFFFFF")
_GROUP_FILL = PatternFill("solid", fgColor="E5E7EB")
_GROUP_FONT = Font(bold=True, color="111827")
_COVER_LABEL_FONT = Font(bold=True, color="111827")

# Column order for the per-trade deliverables sheet — narrower than the
# master register since the trade is implied by the package, and the
# subcontractor only needs the actionable columns.
_TENDER_COLUMNS = [
    "index",
    "service",
    "category",
    "deliverable",
    "applicable_standards",
    "source_document",
    "source_ref",
    "confidence",
    "flags",
]


# --------------------------------------------------------------------------
# Errors + result types
# --------------------------------------------------------------------------


class TenderBuildError(ValueError):
    """Raised for caller-correctable problems (unknown trade, bad format).

    Distinct from generic ``ValueError`` so CLI / API layers can map it to
    a 400-class response without swallowing genuine bugs.
    """


@dataclass(frozen=True)
class TenderBuildResult:
    """Returned by :func:`build_tender_package`.

    ``output_path`` is the file written. The other fields let callers (CLI
    table, API response model) surface the headline numbers without
    re-querying the database.
    """

    output_path: Path
    trade: str
    deliverable_count: int
    service_count: int
    source_document_count: int
    standards_count: int
    fmt: str


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------


def _now_iso() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def _now_compact() -> str:
    return datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")


def _slug(value: str) -> str:
    """Filename-safe slug — same rules as projects._slugify, kept local
    so we don't reach into a private name."""
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower())
    return cleaned.strip("-") or "unnamed"


def tenders_dir_for(project_slug: str) -> Path:
    """Default output directory for a project's tender packages.

    ``<projects_dir>/<slug>.tenders/`` — parallel to ``<slug>.logs/``.
    Created lazily by :func:`build_tender_package`; this helper just
    returns the path so callers can preview it.
    """
    slug = _slug(project_slug)
    return settings.projects_dir / f"{slug}{DEFAULT_TENDERS_SUBDIR}"


def _open_project(project_slug: str) -> tuple[sqlite3.Connection, str]:
    """Open the project SQLite read-only-ish (we only SELECT).

    Returns ``(conn, slug)`` where slug is the on-disk slug derived from
    the supplied name. Raises :class:`FileNotFoundError` if the project
    file is missing — CLI / API map this to a clean error.
    """
    db_path = project_db_path(project_slug)
    if not db_path.exists():
        raise FileNotFoundError(f"Project not found: {db_path}")
    return connect(db_path), db_path.stem


def _project_display_name(conn: sqlite3.Connection, fallback: str) -> str:
    row = conn.execute("SELECT name FROM project LIMIT 1").fetchone()
    if row is None or not row["name"]:
        return fallback
    return str(row["name"])


def _decode_json_list(raw: object) -> list[str]:
    """Parse a JSON list from the schema, defensively. Schema CHECK
    enforces ``json_valid`` so we should not see garbage in practice,
    but exporters must never crash on a single bad row."""
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(x) for x in raw]
    if not isinstance(raw, str):
        return [str(raw)]
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, ValueError):
        return []
    if isinstance(parsed, list):
        return [str(x) for x in parsed]
    return []


# --------------------------------------------------------------------------
# Flag humanisation
# --------------------------------------------------------------------------
#
# Raw flags like ``conflicts_with_source_0d2d26e3-…`` are meaningless to a
# non-technical PM reading the tender pack. We resolve the trailing UUID to
# the human-readable filename(s) it points to so the cover page and per-row
# flags read like English. See ``apps/web/src/components/review/
# flagExplanations.ts`` for the matching UI vocabulary.

# Prefix → ("human label prefix", "table the trailing id lives in")
# - ``conflicts_with_source_<conflict_id>``: the trailing id is a row in
#   ``conflict``; the relevant filename(s) are the OTHER party's
#   source_document.filename (joined via conflict_party + deliverable).
# - ``superseded_by_<source_id>``: trailing id is a source_document.id
#   directly. (Defensive: no current writer emits this on flags, but the
#   schema has a ``source_document.superseded_by_id`` column and prompts
#   reference the concept — handle it so future writers don't regress UX.)
# - ``references_source_<source_id>``: trailing id is a source_document.id.
#   (Same defensive note as above.)
_FLAG_PREFIX_LABELS: tuple[tuple[str, str, str], ...] = (
    ("conflicts_with_source_", "conflicts with", "conflict"),
    ("superseded_by_", "superseded by", "source"),
    ("references_source_", "references", "source"),
)

# Loose UUID matcher — accepts canonical 8-4-4-4-12 hex form; tightens
# the boundary so we do not accidentally split on a flag that merely
# happens to share the prefix (e.g. a hypothetical
# ``conflicts_with_source_review`` would fall through to the raw form).
_UUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)


def _filenames_for_conflict(
    conn: sqlite3.Connection, conflict_id: str
) -> list[str]:
    """Return the distinct source filenames involved in a conflict.

    Joins ``conflict_party`` → ``deliverable`` → ``source_document`` so a
    PM reading "conflicts with: A.pdf, B.pdf" understands which docs
    disagree. Audit-only parties (no deliverable) are skipped — they
    don't carry a tender-relevant source.
    """
    rows = conn.execute(
        """
        SELECT DISTINCT sd.filename AS filename
        FROM conflict_party cp
        JOIN deliverable d      ON d.id = cp.party_id AND cp.party_kind = 'deliverable'
        JOIN source_document sd ON sd.id = d.source_id
        WHERE cp.conflict_id = ?
        ORDER BY sd.filename COLLATE NOCASE
        """,
        (conflict_id,),
    ).fetchall()
    return [str(r["filename"]) for r in rows if r["filename"]]


def _filename_for_source(
    conn: sqlite3.Connection, source_id: str
) -> str | None:
    row = conn.execute(
        "SELECT filename FROM source_document WHERE id = ?",
        (source_id,),
    ).fetchone()
    if row is None or not row["filename"]:
        return None
    return str(row["filename"])


def _humanise_flag(
    conn: sqlite3.Connection,
    raw_flag: str,
    cache: dict[str, str],
) -> str:
    """Resolve a raw flag string to a PM-readable label.

    - ``conflicts_with_source_<uuid>`` → ``conflicts with: <filename(s)>``
    - ``superseded_by_<uuid>``         → ``superseded by: <filename>``
    - ``references_source_<uuid>``     → ``references: <filename>``
    - Anything else (including unrecognised prefixes, or known prefixes
      where the trailing token is not a UUID, or DB lookups that find
      nothing) is returned UNCHANGED. Three-outcome discipline: never
      silently drop a flag the PM might need to chase.

    ``cache`` is shared per build-call to dedupe DB hits when the same
    flag appears on many rows (the conflict-flag pattern is the common
    case here).
    """
    if raw_flag in cache:
        return cache[raw_flag]

    for prefix, label, kind in _FLAG_PREFIX_LABELS:
        if not raw_flag.startswith(prefix):
            continue
        trailing = raw_flag[len(prefix) :]
        if not _UUID_RE.match(trailing):
            # Known prefix but non-UUID payload — pass through verbatim.
            break
        try:
            if kind == "conflict":
                names = _filenames_for_conflict(conn, trailing)
                if not names:
                    break
                resolved = f"{label}: " + ", ".join(names)
            else:  # "source"
                name = _filename_for_source(conn, trailing)
                if not name:
                    break
                resolved = f"{label}: {name}"
        except sqlite3.Error:
            # Defensive: never let an export crash on a flag lookup.
            break
        cache[raw_flag] = resolved
        return resolved

    cache[raw_flag] = raw_flag
    return raw_flag


def _decode_source_ref(raw: object) -> dict[str, object]:
    if isinstance(raw, dict):
        return dict(raw)
    if not isinstance(raw, str):
        return {}
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, ValueError):
        return {"rendered": raw}
    if isinstance(parsed, dict):
        return parsed
    return {"rendered": str(parsed)}


# --------------------------------------------------------------------------
# Queries
# --------------------------------------------------------------------------

# We hit ``v_master_register`` rather than ``deliverable`` directly so that
# the "only accepted rows" rule matches whatever the master register view
# defines as accepted (currently: auto_approved + user_accepted +
# user_edited + user_promoted, with superseded parents excluded).
_TRADE_ROWS_SQL = """
SELECT
    d.id                                AS id,
    d.source_ref                        AS source_ref,
    d.applicable_standards              AS applicable_standards,
    d.confidence                        AS confidence,
    d.flags                             AS flags,
    d.deliverables_summary              AS deliverables_summary,
    d.created_at                        AS created_at,
    d.notes                             AS notes,
    sd.filename                         AS source_filename,
    sd.document_class                   AS document_class,
    sd.document_state                   AS document_state,
    sd.revision                         AS source_revision,
    tt.value                            AS trade_value,
    st.value                            AS service_value,
    ct.value                            AS category_value
FROM v_master_register d
LEFT JOIN source_document   sd ON sd.id = d.source_id
LEFT JOIN trade_taxonomy    tt ON tt.id = d.trade_id
LEFT JOIN service_taxonomy  st ON st.id = d.service_id
LEFT JOIN category_taxonomy ct ON ct.id = d.category_id
WHERE LOWER(COALESCE(tt.value, '')) = LOWER(?)
ORDER BY
    COALESCE(st.value, '~~unset')   COLLATE NOCASE,
    COALESCE(ct.value, '~~unset')   COLLATE NOCASE,
    sd.filename,
    d.created_at,
    d.id
"""

_TRADE_LIST_SQL = """
SELECT
    COALESCE(tt.value, '(unassigned)')  AS trade_value,
    COUNT(*)                            AS deliverable_count
FROM v_master_register d
LEFT JOIN trade_taxonomy tt ON tt.id = d.trade_id
GROUP BY trade_value
ORDER BY trade_value COLLATE NOCASE
"""


def list_trades_with_deliverables(project_slug: str) -> list[tuple[str, int]]:
    """Return ``(trade, count)`` pairs for every trade with at least one
    accepted deliverable. Trades with no taxonomy assignment are reported
    as ``"(unassigned)"`` so the caller can surface them to the user
    (borderline outcome — needs review before tendering).

    Read-only: opens and closes its own connection.
    """
    conn, _slug_on_disk = _open_project(project_slug)
    try:
        rows = conn.execute(_TRADE_LIST_SQL).fetchall()
    finally:
        conn.close()
    return [(str(r["trade_value"]), int(r["deliverable_count"])) for r in rows]


# --------------------------------------------------------------------------
# Row → flat payload
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class _TenderRow:
    """Flattened, display-ready view of one accepted deliverable."""

    id: str
    service: str
    category: str
    deliverable: str
    applicable_standards: list[str]
    source_document: str
    source_ref: str
    confidence: str
    flags: list[str]
    document_class: str
    document_state: str
    notes: str


def _flatten(
    row: sqlite3.Row,
    *,
    conn: sqlite3.Connection,
    flag_cache: dict[str, str],
) -> _TenderRow:
    sref = _decode_source_ref(row["source_ref"])
    rendered = sref.get("rendered")
    if not isinstance(rendered, str) or not rendered.strip():
        # Fall back to a simple "page N" / "section X" reconstruction so
        # the subcontractor still has something to navigate to. This is a
        # display-only courtesy; the structured locator is preserved in
        # the source DB and out of scope for the tender export.
        page = sref.get("page")
        section = sref.get("section")
        bits: list[str] = []
        if section:
            bits.append(f"§{section}")
        if page:
            bits.append(f"p.{page}")
        rendered = " ".join(bits) if bits else ""

    raw_flags = _decode_json_list(row["flags"])
    # Humanise then dedupe-preserve-order: two distinct conflict UUIDs can
    # collapse to the same human label (same pair of docs in conflict),
    # and a PM doesn't need to see "conflicts with: A, B" twice on one row.
    seen: set[str] = set()
    humanised_flags: list[str] = []
    for f in raw_flags:
        label = _humanise_flag(conn, f, flag_cache)
        if label not in seen:
            seen.add(label)
            humanised_flags.append(label)
    return _TenderRow(
        id=str(row["id"]),
        service=str(row["service_value"] or "(unassigned service)"),
        category=str(row["category_value"] or "(unassigned category)"),
        deliverable=str(row["deliverables_summary"] or ""),
        applicable_standards=_decode_json_list(row["applicable_standards"]),
        source_document=str(row["source_filename"] or "(unknown)"),
        source_ref=str(rendered or ""),
        confidence=str(row["confidence"] or ""),
        flags=humanised_flags,
        document_class=str(row["document_class"] or ""),
        document_state=str(row["document_state"] or ""),
        notes=str(row["notes"] or ""),
    )


def _group(rows: Iterable[_TenderRow]) -> dict[str, dict[str, list[_TenderRow]]]:
    """Group rows by ``service`` then by ``category``, preserving order
    within each bucket (the SQL already sorts)."""
    grouped: dict[str, dict[str, list[_TenderRow]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for row in rows:
        grouped[row.service][row.category].append(row)
    return grouped


# --------------------------------------------------------------------------
# Cover-sheet payload (shared between xlsx + md writers)
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class _CoverPayload:
    project_name: str
    project_slug: str
    trade: str
    generated_at: str
    deliverable_count: int
    service_count: int
    source_documents: list[str]
    applicable_standards: list[str]
    flag_summary: dict[str, int]
    unassigned_warnings: list[str]


def _build_cover(
    *,
    project_name: str,
    project_slug: str,
    trade: str,
    flat_rows: list[_TenderRow],
) -> _CoverPayload:
    sources = sorted({r.source_document for r in flat_rows if r.source_document})
    standards: set[str] = set()
    flag_counts: dict[str, int] = defaultdict(int)
    unassigned: list[str] = []
    services: set[str] = set()

    for row in flat_rows:
        services.add(row.service)
        for std in row.applicable_standards:
            if std.strip():
                standards.add(std.strip())
        for flag in row.flags:
            if flag.strip():
                flag_counts[flag.strip()] += 1
        # Three-outcome surfacing: rows missing a service or category get
        # called out on the cover sheet so the head contractor can decide
        # whether to chase the gap before issuing the package.
        if row.service.startswith("(unassigned"):
            unassigned.append(f"{row.id}: missing service mapping")
        if row.category.startswith("(unassigned"):
            unassigned.append(f"{row.id}: missing category")

    return _CoverPayload(
        project_name=project_name,
        project_slug=project_slug,
        trade=trade,
        generated_at=_now_iso(),
        deliverable_count=len(flat_rows),
        service_count=len(services),
        source_documents=sources,
        applicable_standards=sorted(standards, key=str.casefold),
        flag_summary=dict(sorted(flag_counts.items())),
        unassigned_warnings=unassigned,
    )


# --------------------------------------------------------------------------
# XLSX writer
# --------------------------------------------------------------------------


def _write_xlsx(
    *,
    output_path: Path,
    cover: _CoverPayload,
    grouped: dict[str, dict[str, list[_TenderRow]]],
) -> None:
    workbook = Workbook()

    # ---- Cover sheet -----------------------------------------------------
    cover_sheet = workbook.active
    cover_sheet.title = "Cover"
    cover_sheet.append(["Tender Package"])
    cover_sheet["A1"].font = Font(bold=True, size=18, color="111827")
    cover_sheet.row_dimensions[1].height = 28

    rows_pairs: list[tuple[str, str]] = [
        ("Project", cover.project_name),
        ("Project slug", cover.project_slug),
        ("Trade", cover.trade),
        ("Generated (UTC)", cover.generated_at),
        ("Deliverable count", str(cover.deliverable_count)),
        ("Distinct services", str(cover.service_count)),
        ("Source documents", str(len(cover.source_documents))),
        ("Distinct applicable standards", str(len(cover.applicable_standards))),
    ]
    cover_sheet.append([])
    for label, value in rows_pairs:
        cover_sheet.append([label, value])
        cell = cover_sheet.cell(row=cover_sheet.max_row, column=1)
        cell.font = _COVER_LABEL_FONT

    # Source document list
    cover_sheet.append([])
    cover_sheet.append(["Source documents"])
    cover_sheet.cell(row=cover_sheet.max_row, column=1).font = _SUBHEADER_FONT
    cover_sheet.cell(row=cover_sheet.max_row, column=1).fill = _SUBHEADER_FILL
    if cover.source_documents:
        for src in cover.source_documents:
            cover_sheet.append([src])
    else:
        cover_sheet.append(["(none — no accepted deliverables had a source)"])

    # Applicable standards list
    cover_sheet.append([])
    cover_sheet.append(["Applicable standards"])
    cover_sheet.cell(row=cover_sheet.max_row, column=1).font = _SUBHEADER_FONT
    cover_sheet.cell(row=cover_sheet.max_row, column=1).fill = _SUBHEADER_FILL
    if cover.applicable_standards:
        for std in cover.applicable_standards:
            cover_sheet.append([std])
    else:
        cover_sheet.append(["(none cited)"])

    # Flag summary — empty-state messaging when nothing is flagged.
    cover_sheet.append([])
    cover_sheet.append(["Flag summary", "Count"])
    for col in cover_sheet[cover_sheet.max_row]:
        col.font = _SUBHEADER_FONT
        col.fill = _SUBHEADER_FILL
    if cover.flag_summary:
        for flag, count in cover.flag_summary.items():
            cover_sheet.append([flag, count])
    else:
        cover_sheet.append(["(no flags on accepted rows)", ""])

    # Items needing review before issue
    if cover.unassigned_warnings:
        cover_sheet.append([])
        cover_sheet.append(["Review before issue (borderline)"])
        cover_sheet.cell(row=cover_sheet.max_row, column=1).font = _SUBHEADER_FONT
        cover_sheet.cell(row=cover_sheet.max_row, column=1).fill = PatternFill(
            "solid", fgColor="B45309"
        )
        for warning in cover.unassigned_warnings:
            cover_sheet.append([warning])

    cover_sheet.column_dimensions["A"].width = 36
    cover_sheet.column_dimensions["B"].width = 64

    # ---- Deliverables sheet ---------------------------------------------
    sheet = workbook.create_sheet(title="Deliverables")
    sheet.append(_TENDER_COLUMNS)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    index = 0
    for service in sorted(grouped, key=str.casefold):
        # Service group header
        sheet.append([f"Service: {service}"])
        header_cell = sheet.cell(row=sheet.max_row, column=1)
        header_cell.font = _GROUP_FONT
        header_cell.fill = _GROUP_FILL
        sheet.merge_cells(
            start_row=sheet.max_row,
            start_column=1,
            end_row=sheet.max_row,
            end_column=len(_TENDER_COLUMNS),
        )

        for category in sorted(grouped[service], key=str.casefold):
            # Category sub-header
            sheet.append([f"  Category: {category}"])
            sub_cell = sheet.cell(row=sheet.max_row, column=1)
            sub_cell.font = Font(italic=True, color="374151")
            sheet.merge_cells(
                start_row=sheet.max_row,
                start_column=1,
                end_row=sheet.max_row,
                end_column=len(_TENDER_COLUMNS),
            )

            for row in grouped[service][category]:
                index += 1
                sheet.append(
                    [
                        index,
                        row.service,
                        row.category,
                        row.deliverable,
                        ", ".join(row.applicable_standards),
                        row.source_document,
                        row.source_ref,
                        row.confidence,
                        ", ".join(row.flags),
                    ]
                )
                # Wrap the deliverable + standards columns so the row
                # stays readable in print-out.
                deliverable_cell = sheet.cell(
                    row=sheet.max_row, column=_TENDER_COLUMNS.index("deliverable") + 1
                )
                deliverable_cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "index": 6,
        "service": 22,
        "category": 18,
        "deliverable": 70,
        "applicable_standards": 28,
        "source_document": 36,
        "source_ref": 22,
        "confidence": 12,
        "flags": 22,
    }
    for col_index, name in enumerate(_TENDER_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = widths[name]
    sheet.freeze_panes = "A2"

    # Empty-state sheet content if there were no deliverables — the cover
    # already says "0", but leaving the deliverables sheet with only
    # headers is confusing without a hint.
    if cover.deliverable_count == 0:
        sheet.append([])
        sheet.append(
            [
                "No accepted deliverables for this trade.",
                "Check the master register: rows may still be quarantined awaiting review.",
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


# --------------------------------------------------------------------------
# Markdown writer
# --------------------------------------------------------------------------


def _md_escape(value: str) -> str:
    """Escape pipe characters so values do not break a Markdown table."""
    return value.replace("|", "\\|").replace("\n", " ").strip()


def _write_md(
    *,
    output_path: Path,
    cover: _CoverPayload,
    grouped: dict[str, dict[str, list[_TenderRow]]],
) -> None:
    lines: list[str] = []
    lines.append(f"# Tender Package — {cover.trade}")
    lines.append("")
    lines.append(f"**Project:** {cover.project_name}  ")
    lines.append(f"**Project slug:** `{cover.project_slug}`  ")
    lines.append(f"**Generated (UTC):** {cover.generated_at}  ")
    lines.append(f"**Deliverable count:** {cover.deliverable_count}  ")
    lines.append(f"**Distinct services:** {cover.service_count}  ")
    lines.append("")

    lines.append("## Source documents")
    if cover.source_documents:
        for src in cover.source_documents:
            lines.append(f"- {src}")
    else:
        lines.append("_(none — no accepted deliverables had a source)_")
    lines.append("")

    lines.append("## Applicable standards")
    if cover.applicable_standards:
        for std in cover.applicable_standards:
            lines.append(f"- {std}")
    else:
        lines.append("_(none cited)_")
    lines.append("")

    lines.append("## Flag summary")
    if cover.flag_summary:
        lines.append("| Flag | Count |")
        lines.append("|---|---:|")
        for flag, count in cover.flag_summary.items():
            lines.append(f"| {_md_escape(flag)} | {count} |")
    else:
        lines.append("_(no flags on accepted rows)_")
    lines.append("")

    if cover.unassigned_warnings:
        lines.append("## Review before issue (borderline)")
        for warning in cover.unassigned_warnings:
            lines.append(f"- {warning}")
        lines.append("")

    lines.append("## Deliverables")
    if cover.deliverable_count == 0:
        lines.append("")
        lines.append(
            "_No accepted deliverables for this trade. Check the master "
            "register — rows may still be quarantined awaiting review._"
        )
    else:
        for service in sorted(grouped, key=str.casefold):
            lines.append("")
            lines.append(f"### {service}")
            for category in sorted(grouped[service], key=str.casefold):
                lines.append("")
                lines.append(f"#### {category}")
                lines.append("")
                lines.append(
                    "| # | Deliverable | Standards | Source | Ref | Confidence | Flags |"
                )
                lines.append("|---:|---|---|---|---|---|---|")
                for index, row in enumerate(grouped[service][category], start=1):
                    lines.append(
                        "| {idx} | {deliverable} | {standards} | {source} | {ref} | {conf} | {flags} |".format(
                            idx=index,
                            deliverable=_md_escape(row.deliverable),
                            standards=_md_escape(", ".join(row.applicable_standards)),
                            source=_md_escape(row.source_document),
                            ref=_md_escape(row.source_ref),
                            conf=_md_escape(row.confidence),
                            flags=_md_escape(", ".join(row.flags)),
                        )
                    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------


def build_tender_package(
    project_slug: str,
    trade: str,
    *,
    output_dir: Path | None = None,
    fmt: str = "xlsx",
) -> Path:
    """Build a per-trade tender package and return the file path written.

    Parameters
    ----------
    project_slug:
        Project name or slug. Resolved via :func:`meridian.projects.project_db_path`
        (which slugifies internally), so either the human name or the slug
        on disk is accepted.
    trade:
        Trade name, matched case-insensitively against ``trade_taxonomy.value``.
    output_dir:
        Directory to write into. Defaults to ``<projects_dir>/<slug>.tenders/``.
    fmt:
        ``"xlsx"`` (default) or ``"md"``.

    Returns
    -------
    Path
        Absolute path to the written file. The filename is
        ``<slug>__tender__<trade-slug>__<UTC-timestamp>.<ext>``.

    Raises
    ------
    FileNotFoundError
        Project SQLite does not exist.
    TenderBuildError
        Unsupported format, or trade not present in this project's
        register (i.e. zero deliverables and zero taxonomy entry).
        Empty-but-known trades (taxonomy hit, zero accepted rows) still
        produce a package — the cover + deliverables sheet explain the
        empty state. This is deliberate: subcontractors sometimes need
        the explicit "nothing for you on this trade" record.
    """
    fmt_normalised = fmt.lower().strip()
    if fmt_normalised not in SUPPORTED_FORMATS:
        raise TenderBuildError(
            f"Unsupported format {fmt!r}; expected one of {SUPPORTED_FORMATS}"
        )

    conn, slug_on_disk = _open_project(project_slug)
    try:
        project_name = _project_display_name(conn, slug_on_disk)

        # Verify the trade exists in taxonomy. This is the "skip / include
        # / borderline" gate: a totally unknown trade is a *skip* (raise),
        # while a known-but-empty trade is an *include* (write the empty
        # package), and a known-with-rows is the happy path.
        taxonomy_hit = conn.execute(
            "SELECT 1 FROM trade_taxonomy WHERE LOWER(value) = LOWER(?) LIMIT 1",
            (trade,),
        ).fetchone()
        rows = conn.execute(_TRADE_ROWS_SQL, (trade,)).fetchall()

        if not taxonomy_hit and not rows:
            raise TenderBuildError(
                f"Trade {trade!r} not found in this project. "
                f"Run `meridian tender list {project_slug}` to see available trades."
            )

        # Per-build cache for flag → human-readable label so the same
        # ``conflicts_with_source_<uuid>`` appearing on N rows only hits
        # the DB once.
        flag_cache: dict[str, str] = {}
        flat_rows = [_flatten(r, conn=conn, flag_cache=flag_cache) for r in rows]
        cover = _build_cover(
            project_name=project_name,
            project_slug=slug_on_disk,
            trade=trade,
            flat_rows=flat_rows,
        )
        grouped = _group(flat_rows)
    finally:
        conn.close()

    target_dir = (
        output_dir if output_dir is not None else tenders_dir_for(slug_on_disk)
    )
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = (
        f"{slug_on_disk}__tender__{_slug(trade)}__{_now_compact()}.{fmt_normalised}"
    )
    output_path = target_dir / filename

    if fmt_normalised == "xlsx":
        _write_xlsx(output_path=output_path, cover=cover, grouped=grouped)
    else:  # "md" — guarded by SUPPORTED_FORMATS check above
        _write_md(output_path=output_path, cover=cover, grouped=grouped)

    _log.info(
        "tender.build",
        project_slug=slug_on_disk,
        trade=trade,
        fmt=fmt_normalised,
        deliverable_count=cover.deliverable_count,
        service_count=cover.service_count,
        source_count=len(cover.source_documents),
        standards_count=len(cover.applicable_standards),
        unassigned_count=len(cover.unassigned_warnings),
        output_path=str(output_path),
    )
    return output_path


__all__ = [
    "DEFAULT_TENDERS_SUBDIR",
    "OutputFormat",
    "SUPPORTED_FORMATS",
    "TenderBuildError",
    "TenderBuildResult",
    "build_tender_package",
    "list_trades_with_deliverables",
    "tenders_dir_for",
]
